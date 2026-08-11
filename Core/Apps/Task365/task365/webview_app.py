"""
Interface desktop de task365 — version pywebview (HTML/CSS/JS).

Meme tableau de bord du JOUR que la version DearPyGui, pour comparer.
L'interface est en HTML/CSS/JS, affichee dans une fenetre native. Le
JavaScript appelle des methodes Python (classe Api) qui, elles, utilisent
db.py — donc la meme logique que le CLI, la TUI et la version DearPyGui.

Lancement : task365 web   (ou : python -m task365.webview_app)

Dependance : pip install pywebview
  (selon l'OS, un moteur de rendu peut etre requis :
   Linux -> paquet 'python3-gi' + WebKit2GTK, souvent deja present ;
   Windows -> WebView2 ; macOS -> rien de plus.)
"""

from __future__ import annotations

import json
from datetime import date, timedelta

try:
    import webview
except ImportError as exc:
    raise SystemExit(
        "L'interface web desktop nécessite pywebview.\n"
        "Installe-le avec : pip install pywebview\n"
        f"(détail : {exc})"
    )

from . import db


def _num(value) -> float | None:
    """Convertit une valeur venue du JS en float, ou None si vide/invalide."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _next_day(iso: str) -> str:
    """Renvoie le lendemain d'une date ISO (pour 'strictement après')."""
    from datetime import date as _date, timedelta
    return (_date.fromisoformat(iso[:10]) + timedelta(days=1)).isoformat()


# Noms français, pour ne pas dépendre de la locale système (souvent en anglais).
_JOURS_FR = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi",
             "dimanche"]
_MOIS_FR = ["", "janvier", "février", "mars", "avril", "mai", "juin",
            "juillet", "août", "septembre", "octobre", "novembre", "décembre"]


def _date_fr(d) -> str:
    """Formate une date en français : 'dimanche 14 juin 2026'."""
    return f"{_JOURS_FR[d.weekday()]} {d.day} {_MOIS_FR[d.month]} {d.year}"


# Unités par défaut pour les métriques connues (les autres restent sans unité).
_METRIC_UNITS = {
    "poids": "kg", "sommeil": "h", "fc_moy": "bpm", "fc": "bpm",
    "tension": None, "glycemie": "g/L", "humeur": None,
}


def _build_metrics(raw: dict | None) -> dict:
    """
    Transforme un dict {nom: valeur} venu du JS en métriques pour db.add_entry.
    Ignore les valeurs vides/non numériques. Applique une unité connue si dispo.
    """
    out: dict = {}
    if not raw:
        return out
    for name, value in raw.items():
        name = (name or "").strip().lower()
        if not name:
            continue
        v = _num(value)
        if v is None:
            continue
        unit = _METRIC_UNITS.get(name)
        out[name] = (v, unit) if unit else v
    return out


class Api:
    """
    Pont JS -> Python. Chaque methode est appelable depuis le JavaScript via
    window.pywebview.api.<nom>(...). On ne met ici AUCUNE logique metier :
    on delegue a db.py et on renvoie des structures simples (dict/list).
    """

    def __init__(self) -> None:
        self.day = date.today()
        self.window = None  # rempli dans run(), sert aux dialogues de fichier

    # -- compteurs pour la barre laterale --
    def sidebar_counts(self) -> dict:
        """Petits compteurs affiches dans le menu (a faire aujourd'hui, etc.)."""
        from datetime import date as _date
        today = _date.today().isoformat()
        tasks_today = [e for e in db.entries_on_day(today)
                       if e["type"] == "task" and not e["done"]]
        contacts = db.list_entries(type="contact")
        recurs = db.list_recurrences(active_only=True)
        return {
            "day": len(tasks_today),
            "contacts": len(contacts),
            "recur": len(recurs),
        }

    # -- navigation de date --
    def get_day(self) -> dict:
        d = self.day
        return {
            "iso": d.isoformat(),
            "human": _date_fr(d),
            "is_today": d == date.today(),
        }

    def shift_day(self, days: int) -> dict:
        self.day = self.day + timedelta(days=int(days))
        return self.load()

    def go_today(self) -> dict:
        self.day = date.today()
        return self.load()

    def go_to_date(self, iso: str) -> dict:
        """Va directement a une date choisie (format AAAA-MM-JJ)."""
        try:
            self.day = date.fromisoformat(iso[:10])
        except (ValueError, TypeError):
            pass  # date invalide : on ne bouge pas
        return self.load()

    # -- action --
    def toggle_task(self, task_id: int, done: bool) -> bool:
        db.set_done(int(task_id), bool(done))
        return True

    # -- creation / modification : taches --
    def add_task(self, title: str, due: str | None = None,
                 tags: str | None = None, freq: str | None = None,
                 interval: int | None = None, priority: bool = False) -> dict:
        """
        Cree une tache simple, OU une recurrence si freq est fourni
        (daily/weekly/monthly). Dans ce cas la date sert de date de debut
        et les taches dues sont generees immediatement.
        """
        from .utils import parse_date, parse_tags
        title = (title or "").strip() or "Sans titre"
        start = parse_date(due) if due else self.day.isoformat()
        if freq:  # -> recurrence
            db.add_recurrence(
                title=title, freq=freq,
                interval=int(interval) if interval else 1,
                start=start, tags=parse_tags(tags) or None, gen_type="task",
            )
            db.run_recurrences()  # genere tout de suite les occurrences dues
        else:     # -> tache simple
            db.add_entry(type="task", title=title, date=start,
                         tags=parse_tags(tags) or None,
                         priority=1 if priority else 0)
        return self.load()

    def update_task(self, task_id: int, title: str | None = None,
                    due: str | None = None, tags: str | None = None,
                    priority: bool | None = None) -> dict:
        from .utils import parse_date, parse_tags
        db.update_entry(
            int(task_id),
            title=(title or None),
            date=parse_date(due) if due else None,
            tags=parse_tags(tags) if tags is not None else None,
            priority=(1 if priority else 0) if priority is not None else None,
        )
        return self.load()

    def delete_task(self, task_id: int) -> dict:
        db.delete_entry(int(task_id))
        return self.load()

    def get_task(self, task_id: int) -> dict | None:
        e = db.get_entry(int(task_id))
        if not e or e["type"] != "task":
            return None
        return {"id": e["id"], "title": e.get("title") or "",
                "date": e.get("date") or "",
                "tags": " ".join(e["tags"]),
                "priority": bool(e.get("priority"))}

    # -- creation / modification : sport --
    def add_sport(self, activity: str, duration: float | None = None,
                  distance: float | None = None, calories: float | None = None,
                  hr: float | None = None, effort: float | None = None,
                  tags: str | None = None) -> dict:
        from .utils import parse_tags
        db.add_sport(
            activity=(activity or "").strip() or "activité",
            duration_min=_num(duration), distance_km=_num(distance),
            calories=_num(calories), hr_avg=_num(hr), effort=_num(effort),
            date=self.day.isoformat(),
            tags=parse_tags(tags) or None,
        )
        return self.load()

    def update_sport(self, sport_id: int, activity: str | None = None,
                     duration: float | None = None, distance: float | None = None,
                     calories: float | None = None, hr: float | None = None,
                     effort: float | None = None, tags: str | None = None) -> dict:
        from .utils import parse_tags
        db.update_sport(
            int(sport_id), activity=(activity or None),
            duration_min=_num(duration), distance_km=_num(distance),
            calories=_num(calories), hr_avg=_num(hr), effort=_num(effort),
            tags=parse_tags(tags) if tags is not None else None,
        )
        return self.load()

    def delete_sport(self, sport_id: int) -> dict:
        db.delete_sport(int(sport_id))
        return self.load()

    def get_sport(self, sport_id: int) -> dict | None:
        v = db.get_sport(int(sport_id))
        if not v:
            return None
        return {"id": v["id"], "activity": v["activity"],
                "duration": v["duration"], "distance": v["distance"],
                "calories": v["calories"], "hr_avg": v["hr_avg"],
                "effort": v["effort"], "tags": " ".join(v["tags"])}

    # -- budget (lecture) --
    def load_budget(self, since: str | None = None, until: str | None = None,
                    account_id: int | None = None, tag: str | None = None,
                    default_view: bool = True) -> dict:
        """
        Charge la vue budget.
        - Soldes : TOUJOURS reels (a aujourd'hui), independants du filtre.
        - Operations + categories : selon le filtre, ou vue par defaut.
        Vue par defaut (default_view=True et aucun filtre de date) :
          les 10 prochaines operations apres aujourd'hui + toutes celles
          du mois en cours jusqu'a aujourd'hui.
        """
        from datetime import date as _date

        # Soldes reels, toujours a aujourd'hui
        accounts = []
        total = 0.0
        for ac in db.list_accounts():
            bal = db.balance_at(ac["id"])
            total += bal
            accounts.append({"id": ac["id"], "name": ac["name"],
                             "balance": round(bal, 2)})

        acc = int(account_id) if account_id else None
        tag = (tag or None)

        applied = {"since": since, "until": until, "account_id": acc, "tag": tag,
                   "default": False}

        if default_view and not since and not until and not acc and not tag:
            # --- vue par defaut intelligente ---
            today = _date.today()
            month_start = today.replace(day=1).isoformat()
            today_iso = today.isoformat()
            # passe : du 1er du mois a aujourd'hui inclus
            past = db.list_budget_entries(since=month_start, until=today_iso)
            # futur : apres aujourd'hui, on prend les 10 premieres
            future_all = db.list_budget_entries(since=_next_day(today_iso))
            future_all.sort(key=lambda o: (o["date"] or ""))
            future = future_all[:10]
            ops_raw = past + future
            cats = db.category_breakdown(since=month_start, until=today_iso)
            applied["default"] = True
            applied["since"] = month_start
            applied["until"] = "(+10 à venir)"
        else:
            # --- vue filtree ---
            ops_raw = db.list_budget_entries(account_id=acc, since=since,
                                             until=until, tag=tag)
            cats = db.category_breakdown(account_id=acc, since=since, until=until)

        ops = []
        for op in ops_raw:
            ops.append({
                "id": op["id"], "date": (op["date"] or "")[:10],
                "title": op.get("title") or "",
                "amount": round(op["amount"] or 0, 2),
                "tags": [t for t in op["tags"] if t != "virement"],
                "reconciled": bool(op.get("reconciled")),
            })
        ops.sort(key=lambda o: o["date"], reverse=True)

        cat_list = [{"category": c["category"], "total": round(c["total"], 2),
                     "count": c["count"]} for c in cats]

        # somme de la periode affichee (utile sous le tableau)
        period_sum = round(sum(o["amount"] for o in ops), 2)

        return {
            "accounts": accounts,
            "accounts_total": round(total, 2),
            "operations": ops,
            "categories": cat_list,
            "applied": applied,
            "period_sum": period_sum,
        }

    # -- budget (ecriture) --
    def add_account(self, name: str, balance: float | None = None,
                    opened: str | None = None) -> dict:
        from .utils import parse_date
        db.add_account((name or "").strip() or "Compte",
                       initial_balance=_num(balance) or 0.0,
                       opened_on=parse_date(opened) if opened else None)
        return self.load_budget()

    def delete_account(self, account_id: int) -> dict:
        db.delete_account(int(account_id))
        return self.load_budget()

    def add_operation(self, account_id: int, amount: float, title: str,
                      op_date: str | None = None, tags: str | None = None,
                      is_expense: bool = False) -> dict:
        from .utils import parse_date, parse_tags
        amt = _num(amount) or 0.0
        if is_expense and amt > 0:
            amt = -amt
        db.add_budget_entry(int(account_id), amt, (title or "").strip() or "Opération",
                            date=parse_date(op_date) if op_date else None,
                            tags=parse_tags(tags) or None)
        return self.load_budget()

    def add_transfer(self, from_id: int, to_id: int, amount: float,
                     title: str | None = None, op_date: str | None = None,
                     tags: str | None = None) -> dict:
        from .utils import parse_date, parse_tags
        db.add_transfer(int(from_id), int(to_id), _num(amount) or 0.0,
                        title=(title or None),
                        date=parse_date(op_date) if op_date else None,
                        tags=parse_tags(tags) or None)
        return self.load_budget()

    def set_reconciled(self, op_id: int, reconciled: bool) -> dict:
        """Coche/decoche le rapprochement bancaire d'une operation, sans
        recharger tout le budget (juste l'etat)."""
        ok = db.set_reconciled(int(op_id), bool(reconciled))
        return {"ok": ok, "id": int(op_id), "reconciled": bool(reconciled)}

    def get_operation(self, op_id: int) -> dict | None:
        op = db.get_budget_entry(int(op_id))
        if not op:
            return None
        return {
            "id": op["id"], "title": op.get("title") or "",
            "amount": op["amount"] or 0, "date": (op["date"] or "")[:10],
            "account_id": op["account_id"],
            "tags": " ".join(t for t in op["tags"] if t != "virement"),
            "is_transfer": db.is_transfer(int(op_id)),
        }

    def update_operation(self, op_id: int, amount: float | None = None,
                         title: str | None = None, op_date: str | None = None,
                         account_id: int | None = None, tags: str | None = None,
                         is_expense: bool = False) -> dict:
        from .utils import parse_date, parse_tags
        if db.is_transfer(int(op_id)):
            # pour un virement : seuls montant et date, appliques aux 2 jambes
            db.update_transfer(int(op_id), amount=_num(amount),
                               date=parse_date(op_date) if op_date else None)
        else:
            amt = _num(amount)
            if amt is not None and is_expense and amt > 0:
                amt = -amt
            acc = int(account_id) if account_id else None
            db.update_budget_entry(int(op_id), amount=amt, title=(title or None),
                                   date=parse_date(op_date) if op_date else None,
                                   account_id=acc,
                                   tags=parse_tags(tags) if tags is not None else None)
        return self.load_budget()

    def delete_operation(self, op_id: int) -> dict:
        db.delete_budget_entry(int(op_id))
        return self.load_budget()

    # -- vue annuelle (graphe type GitHub) --
    def load_year(self, type: str | None = None) -> dict:
        """
        Donnees pour le graphe de contributions sur ~1 an glissant.
        Renvoie les jours (du lundi <= il y a 53 semaines, jusqu'a aujourd'hui),
        chacun avec sa date, son compte d'activite et un niveau 0-4.
        """
        from datetime import date as _date, timedelta
        today = _date.today()
        start = today - timedelta(days=371)
        start -= timedelta(days=start.weekday())  # recule au lundi
        counts = db.activity_counts(start.isoformat(), today.isoformat(),
                                    type=(type or None))
        maxc = max(counts.values()) if counts else 0

        def level(n: int) -> int:
            if n <= 0:
                return 0
            if maxc <= 1:
                return 4
            r = n / maxc
            if r <= 0.25:
                return 1
            if r <= 0.5:
                return 2
            if r <= 0.75:
                return 3
            return 4

        days = []
        d = start
        while d <= today:
            iso = d.isoformat()
            n = counts.get(iso, 0)
            days.append({"date": iso, "count": n, "level": level(n),
                         "weekday": d.weekday(), "day": d.day, "month": d.month})
            d += timedelta(days=1)

        return {"days": days, "total": sum(counts.values()), "max": maxc,
                "start": start.isoformat(), "end": today.isoformat()}

    def load_metric_series(self, name: str) -> dict:
        """
        Serie temporelle d'une metrique (ex: poids, sommeil) sur ~1 an, pour
        tracer une courbe de suivi. Renvoie les points {date, value} dans la
        periode, plus quelques stats utiles (min, max, moyenne, dernier).
        """
        from datetime import date as _date, timedelta
        today = _date.today()
        start = (today - timedelta(days=365)).isoformat()
        points = []
        unit = ""
        for r in db.metric_history((name or "").strip().lower()):
            day = (r.get("when_") or "")[:10]
            if not day or day < start:
                continue
            unit = r.get("unit") or unit
            points.append({"date": day, "value": r["value"]})
        points.sort(key=lambda p: p["date"])
        stats = {}
        if points:
            vals = [p["value"] for p in points]
            stats = {"min": round(min(vals), 2), "max": round(max(vals), 2),
                     "avg": round(sum(vals) / len(vals), 2),
                     "last": points[-1]["value"], "count": len(points)}
        return {"name": name, "unit": unit, "points": points, "stats": stats}

    # -- recurrences --
    def load_recurrences(self) -> list[dict]:
        accounts = {a["id"]: a["name"] for a in db.list_accounts()}
        out = []
        for r in db.list_recurrences(active_only=False):
            out.append({
                "id": r["id"], "title": r["title"], "freq": r["freq"],
                "interval": r["interval"], "start": r["start"],
                "last_run": r.get("last_run") or "", "active": bool(r["active"]),
                "type": r["gen_type"], "tags": r["tags"],
                "account_id": r.get("account_id"),
                "account": accounts.get(r.get("account_id"), ""),
                "amount": r.get("amount"),
            })
        return out

    def add_recurrence(self, title: str, freq: str, interval: int | None = None,
                       start: str | None = None, tags: str | None = None,
                       gen_type: str = "task", account_id: int | None = None,
                       amount: float | None = None,
                       is_expense: bool = False) -> list[dict]:
        """
        Cree une recurrence et genere immediatement les occurrences dues.
        gen_type='task' (defaut) ou 'budget'. Pour un budget, account_id et
        amount sont requis ; is_expense force le montant en negatif.
        freq = daily/weekly/monthly.
        """
        from .utils import parse_date, parse_tags
        from datetime import date as _date
        amt = _num(amount)
        if gen_type == "budget" and amt is not None and is_expense and amt > 0:
            amt = -amt
        db.add_recurrence(
            title=(title or "").strip() or "Sans titre",
            freq=freq if freq in ("daily", "weekly", "monthly") else "weekly",
            interval=int(interval) if interval else 1,
            start=parse_date(start) if start else _date.today().isoformat(),
            tags=parse_tags(tags) or None,
            gen_type=("budget" if gen_type == "budget" else "task"),
            account_id=int(account_id) if (gen_type == "budget" and account_id) else None,
            amount=amt if gen_type == "budget" else None,
        )
        db.run_recurrences()
        return self.load_recurrences()

    def recurrence_accounts(self) -> list[dict]:
        """Comptes disponibles pour une recurrence de budget (pour le menu)."""
        return [{"id": a["id"], "name": a["name"]} for a in db.list_accounts()]

    def generate_recurrences_until(self, until: str | None = None) -> dict:
        """
        Genere a l'avance les occurrences dues jusqu'a la date 'until' (incluse).
        Idempotent : ne recree jamais ce qui existe deja. Renvoie le nombre
        d'occurrences creees et la liste des recurrences a jour.
        """
        from .utils import parse_date
        from datetime import date as _date
        target = parse_date(until) if until else _date.today().isoformat()
        created = db.run_recurrences(until_iso=target)
        return {"count": len(created), "until": target,
                "recurrences": self.load_recurrences()}

    def toggle_recurrence(self, rec_id: int, active: bool) -> list[dict]:
        db.set_recurrence_active(int(rec_id), bool(active))
        db.run_recurrences()
        return self.load_recurrences()

    def delete_recurrence(self, rec_id: int) -> list[dict]:
        db.delete_recurrence(int(rec_id))
        return self.load_recurrences()

    # -- contacts & notes --
    def load_contacts(self) -> list[dict]:
        """Liste des contacts triés alphabétiquement, avec leur nb de notes."""
        contacts = db.list_entries(type="contact")
        out = []
        for c in contacts:
            notes = db.list_entries(type="note", parent_id=c["id"])
            out.append({
                "id": c["id"], "name": c.get("title") or "",
                "info": c.get("body") or "", "tags": c["tags"],
                "note_count": len(notes),
            })
        out.sort(key=lambda x: x["name"].lower())
        return out

    def get_contact(self, contact_id: int) -> dict | None:
        c = db.get_entry(int(contact_id))
        if not c or c["type"] != "contact":
            return None
        notes = db.list_entries(type="note", parent_id=c["id"])
        notes_out = [{"id": n["id"], "body": n.get("body") or "",
                      "date": (n.get("date") or n["created_at"] or "")[:10]}
                     for n in notes]
        notes_out.sort(key=lambda n: n["date"], reverse=True)
        return {"id": c["id"], "name": c.get("title") or "",
                "info": c.get("body") or "", "tags": " ".join(c["tags"]),
                "notes": notes_out}

    def add_contact(self, name: str, info: str | None = None,
                    tags: str | None = None) -> list[dict]:
        from .utils import parse_tags
        db.add_entry(type="contact", title=(name or "").strip() or "Sans nom",
                     body=(info or None), tags=parse_tags(tags) or None)
        return self.load_contacts()

    def update_contact(self, contact_id: int, name: str | None = None,
                       info: str | None = None, tags: str | None = None) -> dict | None:
        from .utils import parse_tags
        db.update_entry(int(contact_id), title=(name or None),
                        body=(info if info is not None else None),
                        tags=parse_tags(tags) if tags is not None else None)
        return self.get_contact(int(contact_id))

    def delete_contact(self, contact_id: int) -> list[dict]:
        # supprime le contact ET ses notes (cascade via parent_id)
        db.delete_entry(int(contact_id))
        return self.load_contacts()

    def add_note(self, contact_id: int, body: str,
                 note_date: str | None = None) -> dict | None:
        from .utils import parse_date
        from datetime import date as _date
        c = db.get_entry(int(contact_id))
        if not c or c["type"] != "contact":
            return None
        db.add_entry(type="note", body=(body or "").strip(),
                     date=parse_date(note_date) if note_date else _date.today().isoformat(),
                     parent_id=int(contact_id))
        return self.get_contact(int(contact_id))

    def update_note(self, note_id: int, body: str | None = None,
                    note_date: str | None = None) -> dict | None:
        from .utils import parse_date
        n = db.get_entry(int(note_id))
        if not n or n["type"] != "note":
            return None
        parent = n["parent_id"]
        db.update_entry(int(note_id), body=(body if body is not None else None),
                        date=parse_date(note_date) if note_date else None)
        return self.get_contact(parent) if parent else None

    def delete_note(self, note_id: int) -> dict | None:
        n = db.get_entry(int(note_id))
        if not n or n["type"] != "note":
            return None
        parent = n["parent_id"]
        db.delete_entry(int(note_id))
        return self.get_contact(parent) if parent else None

    # -- MODULE DEPENSES --
    def list_tax_codes(self) -> list[dict]:
        return db.list_tax_codes()

    def get_tax_code(self, code_id: int) -> dict | None:
        return db.get_tax_code(int(code_id))

    def default_tax_code(self) -> dict | None:
        return db.default_tax_code()

    def add_tax_code(self, name: str, lines: list | None = None,
                     is_default: bool = False) -> list[dict]:
        parsed = [(l.get("name", ""), _num(l.get("rate")) or 0)
                  for l in (lines or []) if (l.get("name") or "").strip()]
        db.add_tax_code((name or "").strip() or "Code", parsed, bool(is_default))
        return db.list_tax_codes()

    def update_tax_code(self, code_id: int, name: str | None = None,
                        lines: list | None = None,
                        is_default: bool | None = None) -> list[dict]:
        parsed = None
        if lines is not None:
            parsed = [(l.get("name", ""), _num(l.get("rate")) or 0)
                      for l in lines if (l.get("name") or "").strip()]
        db.update_tax_code(int(code_id), name=(name or None), lines=parsed,
                           is_default=is_default)
        return db.list_tax_codes()

    def delete_tax_code(self, code_id: int) -> list[dict]:
        db.delete_tax_code(int(code_id))
        return db.list_tax_codes()

    def list_expense_categories(self) -> list[dict]:
        return db.list_expense_categories()

    def add_expense_category(self, name: str) -> list[dict]:
        if (name or "").strip():
            db.add_expense_category(name.strip())
        return db.list_expense_categories()

    def update_expense_category(self, cat_id: int, name: str) -> list[dict]:
        if (name or "").strip():
            db.update_expense_category(int(cat_id), name.strip())
        return db.list_expense_categories()

    def delete_expense_category(self, cat_id: int) -> list[dict]:
        db.delete_expense_category(int(cat_id))
        return db.list_expense_categories()

    def compute_taxes(self, amount_ht, amount_ttc, tax_code_id,
                      no_tax: bool = False, last_edited: str = "ht") -> dict:
        """Calcul en direct pour le formulaire (pas d'enregistrement)."""
        code = db.get_tax_code(int(tax_code_id)) if tax_code_id else None
        return db.compute_taxes(_num(amount_ht), _num(amount_ttc), code,
                                bool(no_tax), last_edited or "ht")

    def load_expenses(self, since: str | None = None, until: str | None = None,
                      category_id: int | None = None,
                      op_type: str | None = None) -> dict:
        cats = {c["id"]: c["name"] for c in db.list_expense_categories()}
        codes = {c["id"]: c["name"] for c in db.list_tax_codes()}
        rows = db.list_expenses(
            since=since, until=until,
            category_id=int(category_id) if category_id else None,
            op_type=op_type if op_type in ("expense", "income") else None)
        items = []
        # totaux separes : depenses (taxes payees) vs revenus (taxes percues)
        agg = {"expense": {"ht": 0.0, "ttc": 0.0, "tip": 0.0, "tax": 0.0,
                           "tax_by_name": {}},
               "income": {"ht": 0.0, "ttc": 0.0, "tip": 0.0, "tax": 0.0,
                          "tax_by_name": {}}}
        for e in rows:
            t = e["op_type"] if e["op_type"] in agg else "expense"
            a = agg[t]
            a["ht"] += e["amount_ht"] or 0
            a["ttc"] += e["amount_ttc"] or 0
            a["tip"] += e["tip"] or 0
            a["tax"] += e["tax_total"] or 0
            for tname, tamt in (e["tax_detail"] or {}).items():
                a["tax_by_name"][tname] = round(
                    a["tax_by_name"].get(tname, 0.0) + (tamt or 0), 2)
            items.append({
                "id": e["id"], "date": (e["date"] or "")[:10],
                "label": e["label"] or "", "op_type": e["op_type"],
                "third_party": e["third_party"],
                "category": cats.get(e["category_id"], ""),
                "category_id": e["category_id"],
                "tax_code": codes.get(e["tax_code_id"], ""),
                "tax_code_id": e["tax_code_id"],
                "no_tax": e["no_tax"], "amount_ht": round(e["amount_ht"], 2),
                "amount_ttc": round(e["amount_ttc"], 2),
                "tax_total": round(e["tax_total"], 2),
                "tax_detail": e["tax_detail"], "tip": round(e["tip"], 2),
                "currency": e["currency"],
                "amount_currency": e["amount_currency"],
                "is_personal": e["is_personal"],
            })

        def _round(block):
            return {"ht": round(block["ht"], 2), "ttc": round(block["ttc"], 2),
                    "tip": round(block["tip"], 2), "tax": round(block["tax"], 2),
                    "tax_by_name": block["tax_by_name"],
                    "paid": round(block["ttc"] + block["tip"], 2)}

        exp = _round(agg["expense"])
        inc = _round(agg["income"])
        # solde de taxes a remettre = taxes percues (revenus) - taxes payees (depenses)
        net_tax = round(inc["tax"] - exp["tax"], 2)

        return {
            "expenses": items,
            "categories": db.list_expense_categories(),
            "tax_codes": db.list_tax_codes(),
            "applied": {"since": since, "until": until,
                        "category_id": int(category_id) if category_id else None,
                        "op_type": op_type if op_type in ("expense", "income") else None,
                        "count": len(items)},
            "totals_expense": exp,
            "totals_income": inc,
            "net_tax": net_tax,
            # compat : anciens totaux globaux (depenses)
            "totals": {"ht": exp["ht"], "ttc": exp["ttc"], "tip": exp["tip"],
                       "tax_total": exp["tax"], "tax_by_name": exp["tax_by_name"],
                       "paid": exp["paid"]},
        }

    def get_expense(self, expense_id: int) -> dict | None:
        e = db.get_expense(int(expense_id))
        if not e:
            return None
        e["date"] = (e["date"] or "")[:10]
        return e

    def save_expense(self, data: dict) -> dict:
        """Cree ou met a jour une depense a partir du formulaire."""
        from .utils import parse_date
        from datetime import date as _date
        d = data or {}
        code_id = d.get("tax_code_id")
        code = db.get_tax_code(int(code_id)) if code_id else None
        no_tax = bool(d.get("no_tax"))
        last = d.get("last_edited") or "ht"
        calc = db.compute_taxes(_num(d.get("amount_ht")), _num(d.get("amount_ttc")),
                                code, no_tax, last)
        date = parse_date(d.get("date")) if d.get("date") else _date.today().isoformat()
        fields = dict(
            date=date, label=(d.get("label") or None),
            category_id=int(d["category_id"]) if d.get("category_id") else None,
            tax_code_id=int(code_id) if code_id else None, no_tax=no_tax,
            amount_ht=calc["amount_ht"], amount_ttc=calc["amount_ttc"],
            tax_total=calc["tax_total"], tax_detail=calc["detail"],
            tip=_num(d.get("tip")) or 0.0,
            currency=(d.get("currency") or "CAD").strip() or "CAD",
            amount_currency=_num(d.get("amount_currency")),
            is_personal=bool(d.get("is_personal", True)),
            op_type=("income" if d.get("op_type") == "income" else "expense"),
            third_party=(d.get("third_party") or None),
        )
        if d.get("id"):
            db.update_expense(int(d["id"]), **fields)
        else:
            db.add_expense(**fields)
        return self.load_expenses()

    def delete_expense(self, expense_id: int) -> dict:
        db.delete_expense(int(expense_id))
        return self.load_expenses()

    def export_expenses(self, since: str | None = None, until: str | None = None,
                        category_id: int | None = None) -> dict:
        """
        Exporte les depenses (filtrees) en .xlsx, via une boite de dialogue
        native 'Enregistrer sous'. Compatible Excel / OnlyOffice / LibreOffice.
        Renvoie {ok, path|message}.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {"ok": False,
                    "message": "Le module openpyxl est requis. Installe-le avec : "
                               "pipx inject task365 openpyxl"}

        data = self.load_expenses(since, until,
                                  int(category_id) if category_id else None)
        rows = data["expenses"]
        totals = data["totals"]

        # nom de fichier propose
        suffix = ""
        if since or until:
            suffix = f"_{(since or 'debut')}_a_{(until or 'fin')}"
        default_name = f"depenses{suffix}.xlsx"

        # boite de dialogue native 'enregistrer sous'
        try:
            win = self.window
            if win is None and getattr(webview, "windows", None):
                win = webview.windows[0]
            result = win.create_file_dialog(
                webview.SAVE_DIALOG, save_filename=default_name,
                file_types=("Classeur Excel (*.xlsx)",))
        except Exception as exc:
            return {"ok": False, "message": f"Dialogue indisponible : {exc}"}
        if not result:
            return {"ok": False, "message": "Export annulé."}
        path = result if isinstance(result, str) else result[0]
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # construit le classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dépenses"
        # noms de taxes presents (colonnes dynamiques : TPS, TVQ, TVA...)
        tax_names = list(totals.get("tax_by_name", {}).keys())
        headers = (["Date", "Type", "Description", "Tiers", "Catégorie", "Code taxe", "Compte", "HT (CAD)"]
                   + [f"{t} (CAD)" for t in tax_names]
                   + ["Total taxes", "TTC (CAD)", "Pourboire", "Payé",
                      "Devise", "Montant devise"])
        ws.append(headers)
        head_fill = PatternFill("solid", fgColor="1F3864")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")

        for e in rows:
            detail = e.get("tax_detail") or {}
            line = ([e["date"],
                     ("Revenu" if e["op_type"] == "income" else "Dépense"),
                     e["label"], e.get("third_party") or "", e["category"],
                     ("sans taxe" if e["no_tax"] else e["tax_code"]),
                     ("perso" if e["is_personal"] else "entreprise"),
                     e["amount_ht"]]
                    + [round(detail.get(t, 0), 2) for t in tax_names]
                    + [e["tax_total"], e["amount_ttc"], e["tip"],
                       round(e["amount_ttc"] + e["tip"], 2),
                       e["currency"], e.get("amount_currency") or ""])
            ws.append(line)

        # ligne de totaux
        ws.append([])
        total_line = (["TOTAUX", "", "", "", "", "", "", totals["ht"]]
                      + [totals["tax_by_name"].get(t, 0) for t in tax_names]
                      + [totals["tax_total"], totals["ttc"], totals["tip"],
                         totals["paid"], "", ""])
        ws.append(total_line)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)

        # largeurs de colonnes
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
                max(12, len(str(h)) + 2)

        try:
            wb.save(path)
        except Exception as exc:
            return {"ok": False, "message": f"Échec de l'écriture : {exc}"}
        return {"ok": True, "path": path, "count": len(rows)}

    def export_expenses_csv(self, since: str | None = None,
                            until: str | None = None,
                            category_id: int | None = None) -> dict:
        """
        Exporte les depenses (filtrees) en .csv. Aucune dependance externe.
        Encodage UTF-8 avec BOM pour qu'Excel/OnlyOffice lisent bien les accents.
        Separateur ';' (convention FR/Quebec dans les tableurs).
        """
        import csv

        data = self.load_expenses(since, until,
                                  int(category_id) if category_id else None)
        rows = data["expenses"]
        totals = data["totals"]
        tax_names = list(totals.get("tax_by_name", {}).keys())

        suffix = ""
        if since or until:
            suffix = f"_{(since or 'debut')}_a_{(until or 'fin')}"
        default_name = f"depenses{suffix}.csv"

        try:
            win = self.window
            if win is None and getattr(webview, "windows", None):
                win = webview.windows[0]
            result = win.create_file_dialog(
                webview.SAVE_DIALOG, save_filename=default_name,
                file_types=("Fichier CSV (*.csv)",))
        except Exception as exc:
            return {"ok": False, "message": f"Dialogue indisponible : {exc}"}
        if not result:
            return {"ok": False, "message": "Export annulé."}
        path = result if isinstance(result, str) else result[0]
        if not path.lower().endswith(".csv"):
            path += ".csv"

        headers = (["Date", "Type", "Description", "Tiers", "Catégorie", "Code taxe", "Compte", "HT (CAD)"]
                   + [f"{t} (CAD)" for t in tax_names]
                   + ["Total taxes", "TTC (CAD)", "Pourboire", "Payé",
                      "Devise", "Montant devise"])
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(headers)
                for e in rows:
                    detail = e.get("tax_detail") or {}
                    w.writerow(
                        [e["date"],
                         ("Revenu" if e["op_type"] == "income" else "Dépense"),
                         e["label"], e.get("third_party") or "", e["category"],
                         ("sans taxe" if e["no_tax"] else e["tax_code"]),
                         ("perso" if e["is_personal"] else "entreprise"),
                         e["amount_ht"]]
                        + [round(detail.get(t, 0), 2) for t in tax_names]
                        + [e["tax_total"], e["amount_ttc"], e["tip"],
                           round(e["amount_ttc"] + e["tip"], 2),
                           e["currency"], e.get("amount_currency") or ""])
                w.writerow([])
                w.writerow(
                    ["TOTAUX", "", "", "", "", "", "", totals["ht"]]
                    + [totals["tax_by_name"].get(t, 0) for t in tax_names]
                    + [totals["tax_total"], totals["ttc"], totals["tip"],
                       totals["paid"], "", ""])
        except Exception as exc:
            return {"ok": False, "message": f"Échec de l'écriture : {exc}"}
        return {"ok": True, "path": path, "count": len(rows)}

    # -- creation / modification : journal --
    def add_journal(self, body: str, metrics: dict | None = None,
                    tags: str | None = None) -> dict:
        from .utils import parse_tags
        db.add_entry(type="journal", body=(body or "").strip(),
                     date=self.day.isoformat(),
                     tags=parse_tags(tags) or None,
                     metrics=_build_metrics(metrics) or None)
        return self.load()

    def update_journal(self, entry_id: int, body: str | None = None,
                       metrics: dict | None = None,
                       tags: str | None = None) -> dict:
        from .utils import parse_tags
        # met a jour texte/tags...
        db.update_entry(int(entry_id), body=(body if body is not None else None),
                        tags=parse_tags(tags) if tags is not None else None)
        # ... puis REMPLACE entierement les metriques (le formulaire envoie
        # l'etat final voulu, donc une metrique retiree doit disparaitre).
        db.replace_metrics(int(entry_id), _build_metrics(metrics))
        return self.load()

    def delete_journal(self, entry_id: int) -> dict:
        db.delete_entry(int(entry_id))
        return self.load()

    def get_journal(self, entry_id: int) -> dict | None:
        e = db.get_entry(int(entry_id))
        if not e or e["type"] != "journal":
            return None
        # toutes les métriques, sous forme {nom: valeur}
        metrics = {n: i["value"] for n, i in e["metrics"].items()}
        return {"id": e["id"], "body": e.get("body") or "",
                "metrics": metrics, "tags": " ".join(e["tags"])}

    # -- creation / modification : alimentation --
    def add_food(self, label: str | None = None, qty: str | None = None,
                 kcal: float | None = None, prot: float | None = None,
                 gluc: float | None = None, sat: float | None = None,
                 insat: float | None = None, fibres: float | None = None,
                 meal: str | None = None, tags: str | None = None) -> dict:
        from .utils import parse_tags
        db.add_food(
            label=(label or None), qty=(qty or None),
            kcal=_num(kcal), protein=_num(prot), carbs=_num(gluc),
            fat_sat=_num(sat), fat_unsat=_num(insat), fiber=_num(fibres),
            date=self.day.isoformat(), meal=(meal or None),
            tags=parse_tags(tags) or None,
        )
        return self.load()

    def update_food(self, entry_id: int, label: str | None = None,
                    qty: str | None = None, kcal: float | None = None,
                    prot: float | None = None, gluc: float | None = None,
                    sat: float | None = None, insat: float | None = None,
                    fibres: float | None = None, meal: str | None = None) -> dict:
        from .utils import parse_tags
        # le "repas" est stocke comme tag (coherent avec add_food)
        tags = parse_tags(meal) if meal else None
        db.update_food(
            int(entry_id), label=(label or None), qty=(qty or None),
            kcal=_num(kcal), protein=_num(prot), carbs=_num(gluc),
            fat_sat=_num(sat), fat_unsat=_num(insat), fiber=_num(fibres),
            tags=tags,
        )
        return self.load()

    def delete_food(self, entry_id: int) -> dict:
        db.delete_food(int(entry_id))
        return self.load()

    def get_food(self, entry_id: int) -> dict | None:
        e = db.get_entry(int(entry_id))
        if not e or e["type"] != "food":
            return None
        v = db._food_view(e)
        n = v["nutrients"]
        return {"id": v["id"], "label": v["label"], "qty": v["qty"],
                "kcal": n["kcal"], "prot": n["protein"], "gluc": n["carbs"],
                "sat": n["fat_sat"], "insat": n["fat_unsat"], "fibres": n["fiber"],
                "meal": " ".join(v["tags"]), "tags": " ".join(v["tags"])}

    # -- chargement des donnees du jour --
    def load(self) -> dict:
        db.run_recurrences()
        iso = self.day.isoformat()
        entries = db.entries_on_day(iso)

        tasks = [
            {"id": e["id"], "title": e.get("title") or "",
             "done": bool(e["done"]), "tags": e["tags"],
             "priority": bool(e.get("priority"))}
            for e in entries if e["type"] == "task"
        ]
        # tri : non faites avant faites ; parmi elles, prioritaires d'abord
        tasks.sort(key=lambda t: (t["done"], not t["priority"]))

        journal = []
        for e in entries:
            if e["type"] != "journal":
                continue
            mood = e["metrics"].get("humeur", {}).get("value")
            other = {n: i["value"] for n, i in e["metrics"].items()
                     if n != "humeur"}
            journal.append({"id": e["id"], "body": e.get("body") or "",
                            "mood": mood, "metrics": other, "tags": e["tags"]})

        sport = []
        for a in db.list_sport(since=iso, until=iso):
            sport.append({
                "id": a["id"],
                "activity": a["activity"], "duration": a["duration"],
                "distance": a["distance"], "pace": a["pace"],
                "calories": a["calories"],
            })

        food = []
        for l in db.list_food_log(date=iso):
            n = l["nutrients"]
            food.append({"id": l["id"],
                         "label": l["label"] or "(sans nom)", "qty": l["qty"],
                         "kcal": n["kcal"]})
        food_totals = db.food_day_totals(iso)

        accounts = []
        total = 0.0
        for ac in db.list_accounts():
            bal = db.balance_at(ac["id"], iso)
            total += bal
            accounts.append({"name": ac["name"], "balance": round(bal, 2)})

        return {
            "day": self.get_day(),
            "tasks": tasks,
            "journal": journal,
            "sport": sport,
            "food": food,
            "food_totals": food_totals,
            "accounts": accounts,
            "accounts_total": round(total, 2),
        }


HTML = r"""
<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<style>
  :root {
    --bg:#15161a; --panel:#1d1f25; --border:#2a2d36; --text:#e6e6eb;
    --muted:#8b8d98; --accent:#d4c87a; --green:#78c88c; --red:#dc7878;
    --cyan:#78bed2; --yellow:#dcc878;
  }
  * { box-sizing:border-box; }
  body { margin:0; background:var(--bg); color:var(--text);
    font-family:-apple-system,"Segoe UI",Roboto,sans-serif; font-size:14px; }
  /* Mise en page : sidebar fixe à gauche + zone de contenu */
  .app { display:flex; min-height:100vh; }
  .sidebar { width:210px; flex-shrink:0; background:#101116;
    border-right:1px solid var(--border); padding:14px 0; transition:width .15s ease; }
  .sidebar.collapsed { width:56px; }
  .sidebar.collapsed .brand-text, .sidebar.collapsed .navgroup-title,
  .sidebar.collapsed .navlabel, .sidebar.collapsed .badge { display:none; }
  .sidebar.collapsed .navitem { justify-content:center; padding:10px 0; }
  .brand { display:flex; align-items:center; justify-content:space-between;
    font-weight:700; color:var(--accent); font-size:15px;
    padding:4px 14px 14px 18px; letter-spacing:.02em; }
  .collapse-btn { background:none; border:none; color:var(--muted); cursor:pointer;
    font-size:16px; padding:2px 6px; }
  .collapse-btn:hover { color:var(--accent); }
  .navgroup { margin-bottom:10px; padding-bottom:10px;
    border-bottom:1px solid #1c1e25; }
  .navgroup:last-child { border-bottom:none; }
  .navgroup-title { font-size:11px; text-transform:uppercase; letter-spacing:.06em;
    color:var(--muted); padding:6px 18px 4px; }
  .navitem { display:flex; align-items:center; gap:8px; padding:8px 18px;
    cursor:pointer; color:var(--text); border-left:3px solid transparent;
    font-size:13.5px; }
  .navitem:hover { background:#181a20; }
  .navitem.active { background:#181a20; border-left-color:var(--accent);
    color:var(--accent); font-weight:600; }
  .navlabel { flex:1; }
  .badge { background:#2a2d36; color:var(--muted); font-size:11px;
    border-radius:9px; padding:1px 7px; min-width:18px; text-align:center; }
  .navitem.active .badge { background:var(--accent); color:#15161a; }
  .content { flex:1; min-width:0; }
  header { padding:10px 16px; border-bottom:1px solid var(--border);
    display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
  button { background:var(--panel); color:var(--text); border:1px solid var(--border);
    padding:6px 12px; border-radius:6px; cursor:pointer; font-size:13px; }
  button:hover { border-color:var(--accent); }
  button.primary { background:var(--accent); color:#15161a; border-color:var(--accent); }
  button.small { padding:3px 8px; font-size:12px; }
  button.danger:hover { border-color:var(--red); color:var(--red); }
  #dateLabel { color:var(--cyan); font-weight:600; padding:8px 16px; }
  .cols { display:flex; gap:12px; padding:12px 16px; align-items:flex-start; }
  .col-left { flex:1; min-width:0; } .col-right { width:280px; flex-shrink:0; }
  .panel { background:var(--panel); border:1px solid var(--border);
    border-radius:8px; padding:12px 14px; margin-bottom:12px; }
  .panel h2 { margin:0 0 8px; font-size:13px; text-transform:uppercase;
    letter-spacing:.05em; color:var(--accent);
    display:flex; justify-content:space-between; align-items:center; }
  .muted { color:var(--muted); }
  .row { display:flex; justify-content:space-between; padding:3px 0; align-items:center; }
  .task { display:flex; align-items:center; gap:8px; padding:3px 0; }
  .task.done span.t { color:var(--muted); text-decoration:line-through; }
  .pdot { width:8px; height:8px; border-radius:50%; flex-shrink:0;
    background:transparent; border:1px solid var(--border); }
  .pdot.on { background:var(--red); border-color:var(--red); }
  .tag { color:var(--cyan); font-size:12px; }
  .sep { height:1px; background:var(--border); margin:8px 0; }
  .total { color:var(--yellow); font-weight:600; }
  .pos { color:var(--green); } .neg { color:var(--red); }
  .entry { padding:4px 0; border-bottom:1px solid var(--border); }
  input[type=checkbox] { accent-color:var(--green); width:16px; height:16px; }
  input[type=text], input[type=number], input[type=date], select {
    background:#0f1013; color:var(--text); border:1px solid var(--border);
    border-radius:5px; padding:5px 8px; font-size:13px; }
  input[type=date]::-webkit-calendar-picker-indicator { filter:invert(0.7); cursor:pointer; }
  .datelbl { display:inline-flex; align-items:center; gap:5px; color:var(--muted); font-size:13px; }
  .form { display:flex; flex-wrap:wrap; gap:6px; margin-top:8px; align-items:center; }
  .form input[type=text]{ flex:1; min-width:120px; }
  .form input[type=number]{ width:90px; }
  table { width:100%; border-collapse:collapse; font-size:13px; }
  th { text-align:left; color:var(--muted); font-weight:500; padding:4px 6px;
    border-bottom:1px solid var(--border); }
  td { padding:4px 6px; border-bottom:1px solid var(--border); }
  .hide { display:none; }
  .iconbtn { cursor:pointer; color:var(--muted); padding:0 4px; }
  .iconbtn:hover { color:var(--accent); }
  .ycell { display:inline-block; width:12px; height:12px; border-radius:3px;
    background:#22252c; }
  .yl0 { background:#22252c; } .yl1 { background:#0e4429; }
  .yl2 { background:#006d32; } .yl3 { background:#26a641; } .yl4 { background:#39d353; }
  .ygrid { display:grid; grid-auto-flow:column; grid-template-rows:repeat(7, 13px);
    gap:3px; }
  .ygrid .cell { width:12px; height:12px; border-radius:3px; }
  .ymonths { display:grid; grid-auto-flow:column; gap:3px; color:var(--muted);
    font-size:11px; margin-bottom:4px; padding-right:30px; }
  .ymonths span { white-space:nowrap; overflow:visible; }
</style>
</head>
<body>
  <div class="app">
    <nav class="sidebar" id="sidebar">
      <div class="brand">
        <span class="brand-text">Task365</span>
        <button class="collapse-btn" onclick="toggleSidebar()" title="Replier/déplier">☰</button>
      </div>
      <div class="navgroup">
        <div class="navgroup-title">Tâches</div>
        <div class="navitem active" id="nav-day" onclick="showTab('day')" title="Jour">
          <span>📅</span><span class="navlabel">Jour</span><span class="badge" id="badge-day"></span></div>
        <div class="navitem" id="nav-year" onclick="showTab('year')" title="Récapitulatif année">
          <span>🟩</span><span class="navlabel">Récapitulatif année</span></div>
        <div class="navitem" id="nav-recur" onclick="showTab('recur')" title="Récurrences">
          <span>🔁</span><span class="navlabel">Récurrences</span><span class="badge" id="badge-recur"></span></div>
      </div>
      <div class="navgroup">
        <div class="navgroup-title">Budget personnel</div>
        <div class="navitem" id="nav-budget" onclick="showTab('budget')" title="Budget">
          <span>💰</span><span class="navlabel">Budget</span></div>
      </div>
      <div class="navgroup">
        <div class="navgroup-title">CRM</div>
        <div class="navitem" id="nav-contacts" onclick="showTab('contacts')" title="Contacts et notes">
          <span>👤</span><span class="navlabel">Contacts et notes</span><span class="badge" id="badge-contacts"></span></div>
      </div>
      <div class="navgroup">
        <div class="navgroup-title">Module Travailleur autonome</div>
        <div class="navitem" id="nav-expenses" onclick="showTab('expenses')" title="Gestion revenu et dépense">
          <span>🧾</span><span class="navlabel">Gestion revenu et dépense</span></div>
        <div class="navitem" id="nav-excat" onclick="showTab('excat')" title="Catégories de dépenses">
          <span>🏷️</span><span class="navlabel">Catégories</span></div>
        <div class="navitem" id="nav-taxes" onclick="showTab('taxes')" title="Taxes">
          <span>％</span><span class="navlabel">Taxes</span></div>
      </div>
    </nav>
    <div class="content">

  <!-- VUE JOUR -->
  <div id="view-day">
    <header>
      <button onclick="shift(-1)">&lsaquo; -1j</button>
      <button class="primary" onclick="today()">Aujourd'hui</button>
      <button onclick="shift(1)">+1j &rsaquo;</button>
      <label class="datelbl" style="gap:6px">aller au
        <input type="date" id="nav_date" onchange="gotoDate(this.value)">
      </label>
      <button onclick="reload()">Rafraîchir</button>
    </header>
    <div id="dateLabel"></div>
    <div class="cols">
      <div class="col-left">
        <div class="panel">
          <h2>Tâches <button class="small" onclick="toggleForm('taskForm')">+ ajouter</button></h2>
          <div id="tasks"></div>
          <div id="taskForm" class="form hide">
            <input type="text" id="tf_title" placeholder="Intitulé de la tâche">
            <label class="datelbl">date <input type="date" id="tf_date"></label>
            <input type="text" id="tf_tags" placeholder="tags (ex: maison,urgent)">
            <label class="datelbl"><input type="checkbox" id="tf_priority"> prioritaire</label>
            <span id="recurWrap" class="form" style="margin:0;gap:6px">
              <select id="tf_freq" onchange="document.getElementById('tf_interval').classList.toggle('hide', !this.value)">
                <option value="">Ponctuelle</option>
                <option value="daily">Quotidienne</option>
                <option value="weekly">Hebdomadaire</option>
                <option value="monthly">Mensuelle</option>
              </select>
              <input type="number" id="tf_interval" class="hide" placeholder="tous les N" min="1" step="1" style="width:90px">
            </span>
            <button class="primary small" id="btnTask" onclick="submitTask()">Créer</button>
          </div>
        </div>
        <div class="panel">
          <h2>Journal <button class="small" onclick="toggleForm('journalForm')">+ ajouter</button></h2>
          <div id="journal"></div>
          <div id="journalForm" class="form hide" style="flex-direction:column;align-items:stretch;gap:8px">
            <input type="text" id="jf_body" placeholder="Quoi de neuf aujourd'hui ?">
            <div class="form" style="margin:0">
              <input type="number" id="jf_mood" placeholder="humeur /10" step="any" style="width:110px">
              <input type="number" id="jf_weight" placeholder="poids" step="any" style="width:90px">
              <input type="number" id="jf_sommeil" placeholder="sommeil (h)" step="any" style="width:110px">
              <input type="number" id="jf_tension" placeholder="tension" step="any" style="width:90px">
              <input type="number" id="jf_fc" placeholder="FC (bpm)" step="any" style="width:100px">
              <input type="text" id="jf_tags" placeholder="tags">
            </div>
            <div id="jf_custom"></div>
            <div class="form" style="margin:0">
              <button class="small" onclick="addMetricRow()">+ métrique</button>
              <span style="flex:1"></span>
              <button class="primary small" id="btnJournal" onclick="submitJournal()">Créer</button>
            </div>
          </div>
        </div>
        <div class="panel">
          <h2>Sport <button class="small" onclick="toggleForm('sportForm')">+ ajouter</button></h2>
          <div id="sport"></div>
          <div id="sportForm" class="form hide">
            <input type="text" id="sf_activity" placeholder="course, vélo, muscu…">
            <input type="number" id="sf_duration" placeholder="min" step="any">
            <input type="number" id="sf_distance" placeholder="km" step="any">
            <input type="number" id="sf_calories" placeholder="kcal" step="any">
            <input type="number" id="sf_hr" placeholder="FC" step="any">
            <input type="number" id="sf_effort" placeholder="effort" step="any">
            <button class="primary small" onclick="submitSport()">Créer</button>
          </div>
        </div>
        <div class="panel">
          <h2>Alimentation <button class="small" onclick="toggleForm('foodForm')">+ ajouter</button></h2>
          <div id="food"></div>
          <div id="foodForm" class="form hide">
            <input type="text" id="ff_label" placeholder="Aliment (optionnel)">
            <input type="text" id="ff_qty" placeholder="qté (ex: 150g)" style="width:110px">
            <input type="number" id="ff_kcal" placeholder="kcal" step="any" style="width:80px">
            <input type="number" id="ff_prot" placeholder="prot" step="any" style="width:70px">
            <input type="number" id="ff_gluc" placeholder="gluc" step="any" style="width:70px">
            <input type="number" id="ff_sat" placeholder="sat" step="any" style="width:60px">
            <input type="number" id="ff_insat" placeholder="insat" step="any" style="width:70px">
            <input type="number" id="ff_fibres" placeholder="fibres" step="any" style="width:70px">
            <input type="text" id="ff_meal" placeholder="repas" style="width:90px">
            <button class="primary small" id="btnFood" onclick="submitFood()">Créer</button>
          </div>
        </div>
      </div>
      <div class="col-right">
        <div class="panel"><h2>Comptes &amp; soldes</h2>
          <div class="muted" style="font-size:12px;margin-bottom:6px">à la date affichée</div>
          <div id="accounts"></div></div>
      </div>
    </div>
  </div>

  <!-- VUE BUDGET -->
  <div id="view-budget" class="hide">
    <div class="cols">
      <div class="col-left">
        <div class="panel">
          <h2>Opérations
            <span>
              <button class="small" onclick="toggleForm('opForm')">+ opération</button>
              <button class="small" onclick="toggleForm('transferForm')">+ virement</button>
            </span>
          </h2>
          <div class="form" style="margin-bottom:8px;border-bottom:1px solid var(--border);padding-bottom:8px">
            <label class="datelbl">du <input type="date" id="flt_since"></label>
            <label class="datelbl">au <input type="date" id="flt_until"></label>
            <select id="flt_account"><option value="">tous comptes</option></select>
            <input type="text" id="flt_tag" placeholder="catégorie" style="width:110px">
            <button class="small primary" onclick="applyFilter()">Filtrer</button>
            <button class="small" onclick="resetFilter()">Réinitialiser</button>
            <span id="flt_info" class="muted" style="font-size:12px"></span>
          </div>
          <div id="opForm" class="form hide">
            <select id="of_account"></select>
            <input type="text" id="of_title" placeholder="Libellé">
            <input type="number" id="of_amount" placeholder="montant" step="any" style="width:100px">
            <label class="datelbl">date <input type="date" id="of_date"></label>
            <input type="text" id="of_tags" placeholder="catégorie">
            <label class="datelbl"><input type="checkbox" id="of_expense"> dépense</label>
            <button class="primary small" id="btnOp" onclick="submitOp()">Ajouter</button>
          </div>
          <div id="transferForm" class="form hide">
            <select id="tr_from"></select>
            <span class="muted">→</span>
            <select id="tr_to"></select>
            <input type="number" id="tr_amount" placeholder="montant" step="any" style="width:100px">
            <label class="datelbl">date <input type="date" id="tr_date"></label>
            <button class="primary small" onclick="submitTransfer()">Virer</button>
          </div>
          <table><thead><tr><th>Date</th><th>Libellé</th><th>Catégorie</th><th style="text-align:right">Montant</th><th style="text-align:center" title="Rapproché avec le compte bancaire">Rappr.</th><th></th></tr></thead>
          <tbody id="b_ops"></tbody></table>
        </div>
      </div>
      <div class="col-right">
        <div class="panel">
          <h2>Comptes <button class="small" onclick="toggleForm('accForm')">+ compte</button></h2>
          <div id="accForm" class="form hide">
            <input type="text" id="ac_name" placeholder="Nom du compte">
            <input type="number" id="ac_balance" placeholder="solde initial" step="any" style="width:110px">
            <label class="datelbl">depuis <input type="date" id="ac_opened"></label>
            <button class="primary small" onclick="submitAccount()">Créer</button>
          </div>
          <div id="b_accounts"></div>
        </div>
        <div class="panel"><h2>Par catégorie</h2><div id="b_cats"></div></div>
      </div>
    </div>
  </div>

  <!-- VUE ANNEE -->
  <div id="view-year" class="hide">
    <div style="padding:16px">
      <div class="panel">
        <h2><span id="y_title">Activité sur l'année</span>
          <select id="y_type" onchange="loadYear()">
            <option value="">Toutes activités</option>
            <option value="task">Tâches</option>
            <option value="sport">Sport</option>
            <option value="journal">Journal</option>
            <option value="food">Alimentation</option>
            <option value="budget">Budget</option>
          </select>
        </h2>
        <div id="y_graph" style="overflow-x:auto"></div>
        <div class="form" style="margin-top:10px;justify-content:flex-end;gap:6px;align-items:center">
          <span class="muted" style="font-size:12px">Moins</span>
          <span class="ycell yl0"></span><span class="ycell yl1"></span>
          <span class="ycell yl2"></span><span class="ycell yl3"></span><span class="ycell yl4"></span>
          <span class="muted" style="font-size:12px">Plus</span>
        </div>
      </div>
      <div class="panel">
        <h2><span id="weight_title">Suivi du poids</span></h2>
        <div id="weight_chart"></div>
      </div>
      <div class="panel">
        <h2><span id="sleep_title">Suivi du sommeil</span></h2>
        <div id="sleep_chart"></div>
      </div>
    </div>
  </div>

  <!-- VUE CONTACTS -->
  <div id="view-contacts" class="hide">
    <div class="cols">
      <div class="col-left" style="max-width:340px">
        <div class="panel">
          <h2>Contacts <button class="small" onclick="toggleForm('contactForm')">+ ajouter</button></h2>
          <div id="contactForm" class="form hide" style="flex-direction:column;align-items:stretch;gap:6px">
            <input type="text" id="cf_name" placeholder="Nom">
            <input type="text" id="cf_info" placeholder="Infos (rôle, dispo…)">
            <input type="text" id="cf_tags" placeholder="tags (ex: pro, perso)">
            <div class="form" style="margin:0;justify-content:flex-end">
              <button class="primary small" id="btnContact" onclick="submitContact()">Créer</button>
            </div>
          </div>
          <div id="contactList"></div>
        </div>
      </div>
      <div class="col-right" style="flex:1;width:auto">
        <div class="panel" id="contactDetail">
          <div class="muted">Sélectionne un contact pour voir ses détails et ses notes.</div>
        </div>
      </div>
    </div>
  </div>

  <!-- VUE RECURRENCES -->
  <div id="view-recur" class="hide">
    <div class="cols">
      <div class="col-left">
        <div class="panel"><h2>Récurrences <span>
          <button class="small" onclick="openRecurForm()">+ récurrence</button>
          <button class="small" onclick="toggleForm('genForm')">⏩ Générer à l'avance</button>
        </span></h2>
          <div id="genForm" class="form hide" style="margin-bottom:10px;border-bottom:1px solid var(--border);padding-bottom:10px">
            <span class="muted" style="font-size:13px">Générer les occurrences dues jusqu'au</span>
            <input type="date" id="gen_until">
            <button class="primary small" onclick="generateAhead()">Générer</button>
            <span id="gen_info" class="muted" style="font-size:12px"></span>
          </div>
          <div id="recurForm" class="form hide" style="margin-bottom:10px;border-bottom:1px solid var(--border);padding-bottom:10px">
            <select id="rf_type" onchange="updateRecurType()">
              <option value="task">Tâche</option>
              <option value="budget">Opération budget</option>
            </select>
            <input type="text" id="rf_title" placeholder="Intitulé" style="flex:1;min-width:140px">
            <select id="rf_freq">
              <option value="daily">Quotidienne</option>
              <option value="weekly" selected>Hebdomadaire</option>
              <option value="monthly">Mensuelle</option>
            </select>
            <input type="number" id="rf_interval" placeholder="tous les N" min="1" step="1" value="1" style="width:90px">
            <label class="datelbl">depuis <input type="date" id="rf_start"></label>
            <input type="text" id="rf_tags" placeholder="tags" style="width:110px">
            <span id="rf_budget_fields" class="form hide" style="margin:0;gap:6px">
              <select id="rf_account"></select>
              <input type="number" id="rf_amount" placeholder="montant" step="any" style="width:100px">
              <label class="datelbl"><input type="checkbox" id="rf_expense" checked> dépense</label>
            </span>
            <button class="primary small" onclick="submitRecur()">Créer</button>
          </div>
          <table><thead><tr><th>Intitulé</th><th>Type</th><th style="text-align:right">Montant</th><th>Fréquence</th><th>Depuis</th><th>Dernière génér.</th><th>Tags</th><th>État</th><th></th></tr></thead>
          <tbody id="r_list"></tbody></table>
        </div>
      </div>
    </div>
  </div>

  <!-- VUE GESTION DES DEPENSES -->
  <div id="view-expenses" class="hide">
    <div style="padding:16px">
      <div class="panel">
        <h2>Gestion revenu et dépense <span>
          <button class="small" onclick="openExpenseForm('expense')">+ dépense</button>
          <button class="small" onclick="openExpenseForm('income')">+ revenu</button>
        </span></h2>
        <div class="form" style="margin-bottom:10px;border-bottom:1px solid var(--border);padding-bottom:10px">
          <label class="datelbl">du <input type="date" id="xflt_since"></label>
          <label class="datelbl">au <input type="date" id="xflt_until"></label>
          <select id="xflt_optype">
            <option value="">revenus + dépenses</option>
            <option value="income">revenus</option>
            <option value="expense">dépenses</option>
          </select>
          <select id="xflt_category"></select>
          <button class="small primary" onclick="applyExpenseFilter()">Filtrer</button>
          <button class="small" onclick="resetExpenseFilter()">Réinitialiser</button>
          <span style="flex:1"></span>
          <button class="small" onclick="quarterFilter()">Trimestre courant</button>
          <button class="small" onclick="exportExpenses('xlsx')" title="Exporter en Excel / OnlyOffice">⬇ Excel</button>
          <button class="small" onclick="exportExpenses('csv')" title="Exporter en CSV">⬇ CSV</button>
        </div>
        <div id="expenseFormWrap" class="hide" style="border:1px solid var(--border);border-radius:8px;padding:12px;margin-bottom:12px">
          <div class="form">
            <select id="xf_optype" onchange="updateOpTypeUI()">
              <option value="expense">Dépense</option>
              <option value="income">Revenu</option>
            </select>
            <label class="datelbl">date <input type="date" id="xf_date"></label>
            <input type="text" id="xf_label" placeholder="Description" style="flex:1;min-width:120px">
            <input type="text" id="xf_third" placeholder="Fournisseur" style="width:150px">
            <select id="xf_category"></select>
            <select id="xf_taxcode" onchange="recalcExpense('ht')"></select>
            <label class="datelbl"><input type="checkbox" id="xf_notax" onchange="recalcExpense('ht')"> pas de taxes</label>
            <label class="datelbl"><input type="checkbox" id="xf_personal" checked> compte perso</label>
          </div>
          <div class="form" style="margin-top:8px">
            <label class="datelbl">HT <input type="number" id="xf_ht" step="any" style="width:100px" oninput="recalcExpense('ht')"></label>
            <label class="datelbl" id="lbl_tps">TPS <input type="number" id="xf_tps" step="any" style="width:90px" readonly></label>
            <label class="datelbl" id="lbl_tvq">TVQ <input type="number" id="xf_tvq" step="any" style="width:90px" readonly></label>
            <label class="datelbl">TTC <input type="number" id="xf_ttc" step="any" style="width:100px" oninput="recalcExpense('ttc')"></label>
            <span class="muted" style="font-size:12px">(montants en CAD)</span>
          </div>
          <div class="form" style="margin-top:8px">
            <label class="datelbl">pourboire <input type="number" id="xf_tip" step="any" style="width:90px"></label>
            <label class="datelbl">devise <input type="text" id="xf_currency" value="CAD" style="width:70px"></label>
            <label class="datelbl"><input type="checkbox" id="xf_curcheck" onchange="document.getElementById('xf_curamount').classList.toggle('hide', !this.checked)"> montant devise en plus</label>
            <input type="number" id="xf_curamount" class="hide" step="any" placeholder="montant devise" style="width:120px">
            <span style="flex:1"></span>
            <button class="primary small" id="btnExpense" onclick="submitExpense()">Enregistrer</button>
            <button class="small" onclick="document.getElementById('expenseFormWrap').classList.add('hide')">Annuler</button>
          </div>
        </div>
        <table><thead><tr><th>Date</th><th>Type</th><th>Description</th><th>Tiers</th><th>Catégorie</th><th>Code</th><th>Compte</th>
          <th style="text-align:right">HT</th><th style="text-align:right">Taxes</th>
          <th style="text-align:right">TTC</th><th style="text-align:right">Pourb.</th>
          <th style="text-align:right">Payé</th><th></th></tr></thead>
          <tbody id="x_list"></tbody></table>
        <div id="x_summary" style="margin-top:12px"></div>
      </div>
    </div>
  </div>

  <!-- VUE CATEGORIES DE DEPENSES -->
  <div id="view-excat" class="hide">
    <div style="padding:16px;max-width:480px">
      <div class="panel">
        <h2>Catégories de dépenses</h2>
        <div class="form">
          <input type="text" id="ec_name" placeholder="Nouvelle catégorie" style="flex:1">
          <button class="primary small" onclick="addExCat()">Ajouter</button>
        </div>
        <div id="ec_list" style="margin-top:10px"></div>
      </div>
    </div>
  </div>

  <!-- VUE TAXES -->
  <div id="view-taxes" class="hide">
    <div style="padding:16px">
      <div class="cols">
        <div class="col-left">
          <div class="panel">
            <h2>Codes de taxe</h2>
            <div id="tax_list"></div>
          </div>
        </div>
        <div class="col-right" style="flex:1;width:auto">
          <div class="panel">
            <h2 id="tax_form_title">Nouveau code de taxe</h2>
            <div class="form" style="flex-direction:column;align-items:stretch;gap:8px">
              <div class="form" style="margin:0">
                <input type="text" id="tx_name" placeholder="Nom du code (ex: Québec, France)" style="flex:1">
                <label class="datelbl"><input type="checkbox" id="tx_default"> par défaut</label>
              </div>
              <div class="muted" style="font-size:12px">Taxes du code (nom + taux %)</div>
              <div id="tx_lines"></div>
              <div class="form" style="margin:0">
                <button class="small" onclick="addTaxLineRow()">+ taxe</button>
                <span style="flex:1"></span>
                <button class="small" onclick="resetTaxForm()">Nouveau</button>
                <button class="primary small" id="btnTax" onclick="submitTaxCode()">Enregistrer</button>
              </div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

    </div><!-- /.content -->
  </div><!-- /.app -->

<script>
  function esc(s){ return (s||"").replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c])); }
  let editingTask=null, editingSport=null, editingJournal=null, editingFood=null;
  let curDay=null;

  function toggleSidebar(){
    document.getElementById('sidebar').classList.toggle('collapsed');
  }
  async function refreshBadges(){
    try{
      const c=await window.pywebview.api.sidebar_counts();
      const set=(id,n)=>{ const el=document.getElementById(id);
        if(el){ el.textContent=n>0?n:''; el.style.display=n>0?'':'none'; } };
      set('badge-day',c.day); set('badge-recur',c.recur); set('badge-contacts',c.contacts);
    }catch(e){}
  }

  function showTab(t){
    ['day','year','budget','contacts','recur','expenses','excat','taxes'].forEach(x=>{
      document.getElementById('view-'+x).classList.toggle('hide', x!==t);
      document.getElementById('nav-'+x).classList.toggle('active', x===t);
    });
    if(t==='budget') loadBudget();
    if(t==='recur') loadRecur();
    if(t==='year') loadYear();
    if(t==='contacts') loadContacts();
    if(t==='expenses') loadExpenses();
    if(t==='excat') loadExCats();
    if(t==='taxes') loadTaxes();
  }
  function toggleForm(id){
    const el=document.getElementById(id);
    el.classList.toggle('hide');
    if(el.classList.contains('hide')) return;
    // pre-remplir les dates au jour affiche a l'ouverture (creation)
    const dateDefaults={taskForm:'tf_date', opForm:'of_date', transferForm:'tr_date'};
    if(dateDefaults[id] && curDay && !editingTask && !editingOp){
      const f=document.getElementById(dateDefaults[id]);
      if(f && !f.value) f.value=curDay;
    }
  }

  async function reload(){ render(await window.pywebview.api.load()); }
  async function shift(n){ render(await window.pywebview.api.shift_day(n)); }
  async function today(){ render(await window.pywebview.api.go_today()); }
  async function gotoDate(iso){ if(iso) render(await window.pywebview.api.go_to_date(iso)); }
  async function toggle(id, el){
    await window.pywebview.api.toggle_task(id, el.checked);
    // recharge pour ré-appliquer le tri (les tâches faites descendent en bas,
    // même si elles sont prioritaires)
    reload();
  }

  async function submitTask(){
    const title=document.getElementById('tf_title').value;
    const date=document.getElementById('tf_date').value;
    const tags=document.getElementById('tf_tags').value;
    const freq=document.getElementById('tf_freq').value;
    const interval=document.getElementById('tf_interval').value;
    const priority=document.getElementById('tf_priority').checked;
    if(!title.trim()) return;
    if(editingTask){
      render(await window.pywebview.api.update_task(editingTask,title,date||null,tags,priority));
      editingTask=null;
    } else {
      render(await window.pywebview.api.add_task(title,date||null,tags,freq||null,interval||null,priority));
    }
    ['tf_title','tf_date','tf_tags','tf_interval'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('tf_freq').value='';
    document.getElementById('tf_priority').checked=false;
    document.getElementById('tf_interval').classList.add('hide');
    document.getElementById('btnTask').textContent='Créer';
    document.getElementById('recurWrap').classList.remove('hide');
    document.getElementById('taskForm').classList.add('hide');
  }
  async function editTask(id){
    const t=await window.pywebview.api.get_task(id); if(!t) return;
    editingTask=id;
    document.getElementById('tf_title').value=t.title;
    document.getElementById('tf_date').value=t.date;
    document.getElementById('tf_tags').value=t.tags;
    document.getElementById('tf_priority').checked=t.priority;
    // en édition, la récurrence ne s'applique pas (on modifie une occurrence)
    document.getElementById('recurWrap').classList.add('hide');
    document.getElementById('btnTask').textContent='Enregistrer';
    document.getElementById('taskForm').classList.remove('hide');
  }
  async function delTask(id){ render(await window.pywebview.api.delete_task(id)); }
  async function togglePriority(id, val){
    // bascule rapide de la priorité depuis la pastille
    render(await window.pywebview.api.update_task(id, null, null, null, !!val));
  }

  async function submitSport(){
    const v=k=>document.getElementById('sf_'+k).value;
    const act=v('activity'); if(!act.trim()) return;
    const args=[act, v('duration')||null, v('distance')||null, v('calories')||null, v('hr')||null, v('effort')||null, null];
    if(editingSport){ render(await window.pywebview.api.update_sport(editingSport,...args)); editingSport=null; }
    else render(await window.pywebview.api.add_sport(...args));
    ['activity','duration','distance','calories','hr','effort'].forEach(k=>document.getElementById('sf_'+k).value='');
    document.getElementById('sportForm').classList.add('hide');
  }
  async function editSport(id){
    const s=await window.pywebview.api.get_sport(id); if(!s) return;
    editingSport=id;
    document.getElementById('sf_activity').value=s.activity;
    document.getElementById('sf_duration').value=s.duration||'';
    document.getElementById('sf_distance').value=s.distance||'';
    document.getElementById('sf_calories').value=s.calories||'';
    document.getElementById('sf_hr').value=s.hr_avg||'';
    document.getElementById('sf_effort').value=s.effort||'';
    document.getElementById('sportForm').classList.remove('hide');
  }
  async function delSport(id){ render(await window.pywebview.api.delete_sport(id)); }

  // JOURNAL
  const PREDEF=['mood','weight','sommeil','tension','fc'];
  const PREDEF_NAME={mood:'humeur',weight:'poids',sommeil:'sommeil',tension:'tension',fc:'fc'};
  function addMetricRow(name='', value=''){
    const wrap=document.getElementById('jf_custom');
    const div=document.createElement('div');
    div.className='form'; div.style.margin='0';
    div.innerHTML=`<input type="text" placeholder="nom (ex: glycémie)" class="mname" value="${name?name.replace(/"/g,'&quot;'):''}" style="width:160px">
      <input type="number" placeholder="valeur" step="any" class="mval" value="${value!==''&&value!==null?value:''}" style="width:110px">
      <span class="iconbtn" onclick="this.closest('.form').remove()">🗑</span>`;
    wrap.appendChild(div);
  }
  function collectMetrics(){
    const m={};
    PREDEF.forEach(k=>{ const v=document.getElementById('jf_'+k).value;
      if(v!=='') m[PREDEF_NAME[k]]=v; });
    document.querySelectorAll('#jf_custom .form').forEach(row=>{
      const n=row.querySelector('.mname').value.trim();
      const v=row.querySelector('.mval').value;
      if(n && v!=='') m[n]=v;
    });
    return m;
  }
  async function submitJournal(){
    const body=document.getElementById('jf_body').value;
    const tags=document.getElementById('jf_tags').value;
    const metrics=collectMetrics();
    if(editingJournal){
      render(await window.pywebview.api.update_journal(editingJournal,body,metrics,tags));
      editingJournal=null;
    } else {
      render(await window.pywebview.api.add_journal(body,metrics,tags));
    }
    ['body','mood','weight','sommeil','tension','fc','tags'].forEach(k=>document.getElementById('jf_'+k).value='');
    document.getElementById('jf_custom').innerHTML='';
    document.getElementById('btnJournal').textContent='Créer';
    document.getElementById('journalForm').classList.add('hide');
  }
  async function editJournal(id){
    const e=await window.pywebview.api.get_journal(id); if(!e) return;
    editingJournal=id;
    document.getElementById('jf_body').value=e.body;
    document.getElementById('jf_tags').value=e.tags;
    // remplir les champs prédéfinis, le reste en lignes libres
    ['mood','weight','sommeil','tension','fc'].forEach(k=>document.getElementById('jf_'+k).value='');
    document.getElementById('jf_custom').innerHTML='';
    const predefNames=Object.values(PREDEF_NAME);
    for(const [name,val] of Object.entries(e.metrics||{})){
      const slot=Object.keys(PREDEF_NAME).find(k=>PREDEF_NAME[k]===name);
      if(slot) document.getElementById('jf_'+slot).value=val;
      else addMetricRow(name,val);
    }
    document.getElementById('btnJournal').textContent='Enregistrer';
    document.getElementById('journalForm').classList.remove('hide');
  }
  async function delJournal(id){ render(await window.pywebview.api.delete_journal(id)); }

  // ALIMENTATION
  async function submitFood(){
    const v=k=>document.getElementById('ff_'+k).value;
    const args=[v('label')||null, v('qty')||null, v('kcal')||null, v('prot')||null,
                v('gluc')||null, v('sat')||null, v('insat')||null, v('fibres')||null];
    if(editingFood){
      // en modification : on transmet aussi le repas (ff_meal)
      render(await window.pywebview.api.update_food(editingFood,...args,v('meal')||null));
      editingFood=null;
    } else {
      render(await window.pywebview.api.add_food(...args,v('meal')||null,null));
    }
    ['label','qty','kcal','prot','gluc','sat','insat','fibres','meal'].forEach(k=>document.getElementById('ff_'+k).value='');
    document.getElementById('btnFood').textContent='Créer';
    document.getElementById('foodForm').classList.add('hide');
  }
  async function editFood(id){
    const f=await window.pywebview.api.get_food(id); if(!f) return;
    editingFood=id;
    const set=(k,val)=>document.getElementById('ff_'+k).value=(val===null||val===0?'':val);
    document.getElementById('ff_label').value=f.label||'';
    document.getElementById('ff_qty').value=f.qty||'';
    set('kcal',f.kcal); set('prot',f.prot); set('gluc',f.gluc);
    set('sat',f.sat); set('insat',f.insat); set('fibres',f.fibres);
    document.getElementById('ff_meal').value=f.meal||'';
    document.getElementById('btnFood').textContent='Enregistrer';
    document.getElementById('foodForm').classList.remove('hide');
  }
  async function delFood(id){ render(await window.pywebview.api.delete_food(id)); }

  // ANNEE (graphe type GitHub)
  const MONTHS_FR=['','jan','fév','mar','avr','mai','juin','juil','aoû','sep','oct','nov','déc'];
  async function loadYear(){
    const type=document.getElementById('y_type').value;
    const y=await window.pywebview.api.load_year(type||null);
    document.getElementById('y_title').textContent =
      `${y.total} activité${y.total>1?'s':''} sur l'année`;
    // grille : les jours arrivent déjà du lundi, 7 par colonne
    let cells='';
    y.days.forEach(d=>{
      cells+=`<div class="cell yl${d.level}" title="${d.date} — ${d.count} activité(s)"></div>`;
    });
    const colCount=Math.ceil(y.days.length/7);
    // étiquettes de mois : on étiquette une colonne quand le mois change
    // (on compare avec le dernier mois étiqueté ; le millésime est pris en
    //  compte car on lit le mois réel de chaque colonne, pas juste son numéro)
    let monthLabels=new Array(colCount).fill('');
    let prevMonth=null;
    for(let col=0; col<colCount; col++){
      const firstDay=y.days[col*7];           // 1er jour (lundi) de la colonne
      if(!firstDay) continue;
      const month=parseInt(firstDay.date.slice(5,7),10);  // mois réel via la date
      if(month!==prevMonth){
        monthLabels[col]=MONTHS_FR[month];
        prevMonth=month;
      }
    }
    const months=monthLabels.map(l=>`<span>${l}</span>`).join('');

    document.getElementById('y_graph').innerHTML=
      `<div class="ymonths" style="grid-template-columns:repeat(${colCount}, 12px)">${months}</div>`+
      `<div class="ygrid">${cells}</div>`;
    // charge aussi les courbes de poids et de sommeil
    loadMetricCharts();
  }

  async function loadMetricCharts(){
    const weight=await window.pywebview.api.load_metric_series('poids');
    const sleep=await window.pywebview.api.load_metric_series('sommeil');
    renderMetricChart('weight_chart','weight_title','Suivi du poids', weight, '#78bed2');
    renderMetricChart('sleep_chart','sleep_title','Suivi du sommeil', sleep, '#dcc878');
  }

  function renderMetricChart(elId, titleId, label, data, color){
    const el=document.getElementById(elId);
    const pts=data.points||[];
    const unit=data.unit?(' '+data.unit):'';
    if(pts.length<2){
      el.innerHTML='<div class="muted">Pas assez de données pour tracer une courbe (saisis cette métrique dans le journal).</div>';
      document.getElementById(titleId).textContent=label;
      return;
    }
    const s=data.stats;
    document.getElementById(titleId).innerHTML=
      `${label} <span class="muted" style="font-size:12px;font-weight:400">`+
      `dernier ${s.last}${unit} · min ${s.min} · max ${s.max} · moy ${s.avg} · ${s.count} mesures</span>`;
    // dimensions
    const W=720, H=200, padL=44, padR=12, padT=14, padB=28;
    const innerW=W-padL-padR, innerH=H-padT-padB;
    const vals=pts.map(p=>p.value);
    let vmin=Math.min(...vals), vmax=Math.max(...vals);
    if(vmin===vmax){ vmin-=1; vmax+=1; } // évite division par zéro
    const pad=(vmax-vmin)*0.1; vmin-=pad; vmax+=pad;
    const t0=new Date(pts[0].date).getTime();
    const t1=new Date(pts[pts.length-1].date).getTime();
    const tx=t=> padL + (t1===t0?innerW/2:(t-t0)/(t1-t0)*innerW);
    const ty=v=> padT + innerH - (v-vmin)/(vmax-vmin)*innerH;
    // points
    const P=pts.map(p=>({x:tx(new Date(p.date).getTime()), y:ty(p.value), v:p.value, d:p.date}));
    const line=P.map((p,i)=>(i?'L':'M')+p.x.toFixed(1)+' '+p.y.toFixed(1)).join(' ');
    // aire sous la courbe
    const area=`M${P[0].x.toFixed(1)} ${(padT+innerH).toFixed(1)} `+
      P.map(p=>'L'+p.x.toFixed(1)+' '+p.y.toFixed(1)).join(' ')+
      ` L${P[P.length-1].x.toFixed(1)} ${(padT+innerH).toFixed(1)} Z`;
    // graduations Y (3 niveaux)
    let yticks='';
    for(let i=0;i<=3;i++){
      const v=vmin+(vmax-vmin)*i/3; const y=ty(v);
      yticks+=`<line x1="${padL}" y1="${y.toFixed(1)}" x2="${W-padR}" y2="${y.toFixed(1)}" stroke="#2a2d36" stroke-width="1"/>`;
      yticks+=`<text x="${padL-6}" y="${(y+3).toFixed(1)}" fill="#8b8d98" font-size="10" text-anchor="end">${v.toFixed(1)}</text>`;
    }
    // étiquettes X : première et dernière date
    const fmt=d=>{const x=new Date(d);return x.getDate()+'/'+(x.getMonth()+1);};
    const xlabels=`<text x="${padL}" y="${H-8}" fill="#8b8d98" font-size="10">${fmt(pts[0].date)}</text>`+
      `<text x="${W-padR}" y="${H-8}" fill="#8b8d98" font-size="10" text-anchor="end">${fmt(pts[pts.length-1].date)}</text>`;
    const dots=P.map(p=>`<circle cx="${p.x.toFixed(1)}" cy="${p.y.toFixed(1)}" r="2.5" fill="${color}"><title>${p.d} : ${p.v}${unit}</title></circle>`).join('');
    el.innerHTML=
      `<svg viewBox="0 0 ${W} ${H}" style="width:100%;height:auto">
        ${yticks}
        <path d="${area}" fill="${color}" opacity="0.12"/>
        <path d="${line}" fill="none" stroke="${color}" stroke-width="2"/>
        ${dots}${xlabels}
      </svg>`;
  }
  // CONTACTS
  let editingContact=null, currentContact=null, editingNote=null;
  async function loadContacts(){
    const list=await window.pywebview.api.load_contacts();
    const el=document.getElementById('contactList');
    if(!list.length){ el.innerHTML='<div class="muted">Aucun contact.</div>'; return; }
    let html='', lastInitial='';
    list.forEach(c=>{
      const ini=(c.name[0]||'?').toUpperCase();
      if(ini!==lastInitial){ html+=`<div class="muted" style="margin-top:6px;font-size:12px">${esc(ini)}</div>`; lastInitial=ini; }
      html+=`<div class="row" style="cursor:pointer" onclick="openContact(${c.id})">
        <span>${esc(c.name)}</span>
        <span class="muted" style="font-size:12px">${c.note_count?c.note_count+' note(s)':''}</span></div>`;
    });
    el.innerHTML=html;
  }
  async function submitContact(){
    const name=document.getElementById('cf_name').value;
    const info=document.getElementById('cf_info').value;
    const tags=document.getElementById('cf_tags').value;
    if(!name.trim()) return;
    if(editingContact){
      await window.pywebview.api.update_contact(editingContact,name,info,tags);
      const id=editingContact; editingContact=null;
      await loadContacts(); openContact(id);
    } else {
      await window.pywebview.api.add_contact(name,info,tags);
      await loadContacts();
    }
    ['cf_name','cf_info','cf_tags'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('btnContact').textContent='Créer';
    document.getElementById('contactForm').classList.add('hide');
  }
  function editContactForm(c){
    editingContact=c.id;
    document.getElementById('cf_name').value=c.name;
    document.getElementById('cf_info').value=c.info;
    document.getElementById('cf_tags').value=c.tags;
    document.getElementById('btnContact').textContent='Enregistrer';
    document.getElementById('contactForm').classList.remove('hide');
  }
  async function delContact(id){
    if(!confirm('Supprimer ce contact et toutes ses notes ?')) return;
    await window.pywebview.api.delete_contact(id);
    currentContact=null;
    document.getElementById('contactDetail').innerHTML='<div class="muted">Sélectionne un contact.</div>';
    loadContacts();
  }
  async function openContact(id){
    const c=await window.pywebview.api.get_contact(id); if(!c) return;
    currentContact=c; renderContactDetail(c);
  }
  function renderContactDetail(c){
    const tags=c.tags? c.tags.split(' ').filter(Boolean).map(t=>`<span class="tag">#${esc(t)}</span>`).join(' '):'';
    let h=`<h2>${esc(c.name)}
      <span><span class="iconbtn" onclick='editContactForm(${JSON.stringify(c).replace(/'/g,"&#39;")})'>✎</span>
      <span class="iconbtn" onclick="delContact(${c.id})">🗑</span></span></h2>`;
    if(c.info) h+=`<div>${esc(c.info)}</div>`;
    if(tags) h+=`<div style="margin-top:4px">${tags}</div>`;
    h+='<div class="sep"></div>';
    h+=`<div class="row"><b>Notes</b><button class="small" onclick="toggleNoteForm()">+ note</button></div>`;
    h+=`<div id="noteForm" class="form hide" style="flex-direction:column;align-items:stretch;gap:6px;margin-top:6px">
        <input type="text" id="nf_body" placeholder="Contenu de la note">
        <div class="form" style="margin:0;justify-content:flex-end">
          <label class="datelbl">date <input type="date" id="nf_date"></label>
          <button class="primary small" id="btnNote" onclick="submitNote()">Ajouter</button>
        </div></div>`;
    if(!c.notes.length) h+='<div class="muted" style="margin-top:6px">Aucune note.</div>';
    else h+=c.notes.map(n=>`<div class="entry"><div class="row">
        <span class="muted" style="font-size:12px">${n.date}</span>
        <span><span class="iconbtn" onclick="editNote(${n.id}, this)">✎</span>
        <span class="iconbtn" onclick="delNote(${n.id})">🗑</span></span></div>
        <div data-note="${n.id}">${esc(n.body)}</div></div>`).join('');
    document.getElementById('contactDetail').innerHTML=h;
  }
  function toggleNoteForm(){ document.getElementById('noteForm').classList.toggle('hide');
    if(!document.getElementById('nf_date').value && curDay) document.getElementById('nf_date').value=curDay; }
  async function submitNote(){
    const body=document.getElementById('nf_body').value;
    const date=document.getElementById('nf_date').value;
    if(!body.trim()||!currentContact) return;
    let c;
    if(editingNote){ c=await window.pywebview.api.update_note(editingNote,body,date||null); editingNote=null;
      document.getElementById('btnNote').textContent='Ajouter'; }
    else c=await window.pywebview.api.add_note(currentContact.id,body,date||null);
    document.getElementById('nf_body').value=''; document.getElementById('nf_date').value='';
    if(c){ currentContact=c; renderContactDetail(c); }
  }
  function editNote(id, el){
    const text=el.closest('.entry').querySelector('[data-note]').textContent;
    editingNote=id;
    document.getElementById('noteForm').classList.remove('hide');
    document.getElementById('nf_body').value=text;
    document.getElementById('btnNote').textContent='Enregistrer';
  }
  async function delNote(id){
    const c=await window.pywebview.api.delete_note(id);
    if(c){ currentContact=c; renderContactDetail(c); }
  }

  // ===== MODULE DEPENSES =====
  let expenseData={categories:[],tax_codes:[]}, editingExpense=null, expenseFilter=null;

  function fillSelect(id, items, valueKey, labelKey, placeholder){
    const s=document.getElementById(id);
    s.innerHTML=(placeholder?`<option value="">${placeholder}</option>`:'')
      + items.map(i=>`<option value="${i[valueKey]}">${esc(i[labelKey])}</option>`).join('');
  }

  async function loadExpenses(){
    const f=expenseFilter||{};
    const d=await window.pywebview.api.load_expenses(f.since||null, f.until||null, f.category_id||null, f.op_type||null);
    expenseData=d;
    fillSelect('xflt_category', d.categories,'id','name','toutes catégories');
    if(f.category_id) document.getElementById('xflt_category').value=f.category_id;
    if(f.op_type) document.getElementById('xflt_optype').value=f.op_type;
    const tb=document.getElementById('x_list');
    tb.innerHTML = !d.expenses.length ?
      '<tr><td colspan="13" class="muted">Aucune opération.</td></tr>' :
      d.expenses.map(e=>{
        const taxes=Object.entries(e.tax_detail||{}).map(([k,v])=>`${k} ${v.toFixed(2)}`).join(', ');
        const paid=(e.amount_ttc+e.tip).toFixed(2);
        const cur=(e.currency&&e.currency!=='CAD'&&e.amount_currency)?` <span class="muted">(${e.amount_currency} ${esc(e.currency)})</span>`:'';
        const compte=e.is_personal?'<span class="muted">perso</span>':'<span class="tag">entreprise</span>';
        const type=e.op_type==='income'?'<span class="pos">revenu</span>':'<span class="neg">dépense</span>';
        return `<tr><td>${e.date}</td><td>${type}</td><td>${esc(e.label)}${cur}</td><td>${esc(e.third_party||'')}</td><td>${esc(e.category)}</td>
          <td>${e.no_tax?'<span class="muted">sans taxe</span>':esc(e.tax_code)}</td>
          <td>${compte}</td>
          <td style="text-align:right">${e.amount_ht.toFixed(2)}</td>
          <td style="text-align:right" class="muted" title="${esc(taxes)}">${e.tax_total.toFixed(2)}</td>
          <td style="text-align:right">${e.amount_ttc.toFixed(2)}</td>
          <td style="text-align:right">${e.tip?e.tip.toFixed(2):'-'}</td>
          <td style="text-align:right"><b>${paid}</b></td>
          <td style="text-align:right"><span class="iconbtn" onclick="editExpense(${e.id})" title="Modifier">✎</span><span class="iconbtn" onclick="duplicateExpense(${e.id})" title="Dupliquer">⧉</span><span class="iconbtn" onclick="delExpense(${e.id})" title="Supprimer">🗑</span></td></tr>`;
      }).join('');
    renderExpenseSummary(d);
  }

  function renderExpenseSummary(d){
    const inc=d.totals_income, exp=d.totals_expense, ap=d.applied;
    let period;
    if(ap.since||ap.until) period=`Période : ${ap.since||'…'} → ${ap.until||'…'}`;
    else period='Toutes les opérations';
    period += `  ·  ${ap.count} opération(s)`;
    const taxLines=(block)=>Object.entries(block.tax_by_name||{})
      .map(([n,v])=>`<div class="row"><span>${esc(n)}</span><span>${v.toFixed(2)}</span></div>`).join('');
    const hasInc=inc.ht||Object.keys(inc.tax_by_name).length;
    const hasExp=exp.ht||Object.keys(exp.tax_by_name).length;
    let h=`<div style="display:flex;gap:12px;flex-wrap:wrap;justify-content:flex-end">`;
    // bloc revenus
    h+=`<div style="border:1px solid var(--border);border-radius:8px;padding:12px;min-width:240px">
      <div class="pos" style="font-weight:600;margin-bottom:6px">Revenus (taxes perçues)</div>
      <div class="row"><span>Total HT</span><span><b>${inc.ht.toFixed(2)}</b></span></div>
      ${taxLines(inc)}
      <div class="row"><span>Total taxes perçues</span><span><b>${inc.tax.toFixed(2)}</b></span></div>
      <div class="row"><span>Total TTC</span><span>${inc.ttc.toFixed(2)}</span></div></div>`;
    // bloc dépenses
    h+=`<div style="border:1px solid var(--border);border-radius:8px;padding:12px;min-width:240px">
      <div class="neg" style="font-weight:600;margin-bottom:6px">Dépenses (taxes payées)</div>
      <div class="row"><span>Total HT</span><span><b>${exp.ht.toFixed(2)}</b></span></div>
      ${taxLines(exp)}
      <div class="row"><span>Total taxes payées</span><span><b>${exp.tax.toFixed(2)}</b></span></div>
      <div class="row"><span>Total TTC</span><span>${exp.ttc.toFixed(2)}</span></div>
      <div class="row muted"><span>Pourboires</span><span>${exp.tip.toFixed(2)}</span></div></div>`;
    h+=`</div>`;
    // solde net de taxes
    const net=d.net_tax;
    h+=`<div style="border:1px solid var(--accent);border-radius:8px;padding:12px;margin-top:10px;max-width:520px;margin-left:auto">
      <div class="muted" style="font-size:12px;margin-bottom:4px">${period}</div>
      <div class="row total"><span>Taxes nettes à remettre (perçues − payées)</span>
        <span class="${net>=0?'pos':'neg'}">${net.toFixed(2)} CAD</span></div>
      <div class="muted" style="font-size:11px;margin-top:4px">${net>=0?'Montant à remettre à Revenu Québec/ARC.':'Crédit en ta faveur (remboursement potentiel).'}</div>
    </div>`;
    document.getElementById('x_summary').innerHTML=h;
  }

  function applyExpenseFilter(){
    expenseFilter={
      since:document.getElementById('xflt_since').value,
      until:document.getElementById('xflt_until').value,
      op_type:document.getElementById('xflt_optype').value,
      category_id:document.getElementById('xflt_category').value,
    };
    loadExpenses();
  }
  function resetExpenseFilter(){
    expenseFilter=null;
    ['xflt_since','xflt_until'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('xflt_category').value='';
    document.getElementById('xflt_optype').value='';
    loadExpenses();
  }
  function quarterFilter(){
    // trimestre civil courant (jan-mar, avr-juin, juil-sep, oct-déc)
    const now=new Date();
    const q=Math.floor(now.getMonth()/3);
    const start=new Date(now.getFullYear(), q*3, 1);
    const end=new Date(now.getFullYear(), q*3+3, 0);
    const iso=dt=>dt.toISOString().slice(0,10);
    document.getElementById('xflt_since').value=iso(start);
    document.getElementById('xflt_until').value=iso(end);
    applyExpenseFilter();
  }
  async function exportExpenses(fmt){
    const f=expenseFilter||{};
    const fn = fmt==='csv' ? 'export_expenses_csv' : 'export_expenses';
    const r=await window.pywebview.api[fn](f.since||null, f.until||null, f.category_id||null);
    if(r.ok) alert(`Export réussi : ${r.count} dépense(s)\\n${r.path}`);
    else alert(r.message||'Export impossible.');
  }

  function openExpenseForm(opType){
    editingExpense=null;
    document.getElementById('btnExpense').textContent='Enregistrer';
    document.getElementById('xf_optype').value=opType||'expense';
    fillSelect('xf_category', expenseData.categories,'id','name','(sans catégorie)');
    fillSelect('xf_taxcode', expenseData.tax_codes,'id','name',null);
    const def=expenseData.tax_codes.find(c=>c.is_default);
    if(def) document.getElementById('xf_taxcode').value=def.id;
    ['xf_label','xf_third','xf_ht','xf_tps','xf_tvq','xf_ttc','xf_tip','xf_curamount'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('xf_date').value=curDay||new Date().toISOString().slice(0,10);
    document.getElementById('xf_currency').value='CAD';
    document.getElementById('xf_notax').checked=false;
    document.getElementById('xf_personal').checked=true;
    document.getElementById('xf_curcheck').checked=false;
    document.getElementById('xf_curamount').classList.add('hide');
    updateOpTypeUI();
    updateTaxLabels();
    document.getElementById('expenseFormWrap').classList.remove('hide');
  }

  function updateOpTypeUI(){
    // adapte le libellé du champ tiers selon revenu/dépense
    const t=document.getElementById('xf_optype').value;
    document.getElementById('xf_third').placeholder = t==='income' ? 'Client' : 'Fournisseur';
  }

  function updateTaxLabels(){
    // adapte les libellés TPS/TVQ selon le code choisi (ex: France -> TVA)
    const codeId=document.getElementById('xf_taxcode').value;
    const code=expenseData.tax_codes.find(c=>String(c.id)===String(codeId));
    const lines=(code&&code.lines)||[];
    const l1=document.getElementById('lbl_tps'), l2=document.getElementById('lbl_tvq');
    if(lines[0]){ l1.childNodes[0].textContent=lines[0].name+' '; l1.style.display=''; }
    else l1.style.display='none';
    if(lines[1]){ l2.childNodes[0].textContent=lines[1].name+' '; l2.style.display=''; }
    else l2.style.display='none';
  }

  async function recalcExpense(edited){
    updateTaxLabels();
    const codeId=document.getElementById('xf_taxcode').value;
    const noTax=document.getElementById('xf_notax').checked;
    const ht=document.getElementById('xf_ht').value;
    const ttc=document.getElementById('xf_ttc').value;
    const r=await window.pywebview.api.compute_taxes(ht||null, ttc||null, codeId||null, noTax, edited);
    // remplit selon le sens (le champ édité n'est pas écrasé)
    if(edited==='ht') document.getElementById('xf_ttc').value=r.amount_ttc;
    else document.getElementById('xf_ht').value=r.amount_ht;
    const vals=Object.values(r.detail);
    document.getElementById('xf_tps').value=vals[0]!==undefined?vals[0]:'';
    document.getElementById('xf_tvq').value=vals[1]!==undefined?vals[1]:'';
  }

  async function submitExpense(){
    const g=id=>document.getElementById(id).value;
    const payload={
      id:editingExpense, date:g('xf_date'), label:g('xf_label'),
      op_type:g('xf_optype'), third_party:g('xf_third'),
      category_id:g('xf_category')||null, tax_code_id:g('xf_taxcode')||null,
      no_tax:document.getElementById('xf_notax').checked,
      is_personal:document.getElementById('xf_personal').checked,
      amount_ht:g('xf_ht')||null, amount_ttc:g('xf_ttc')||null,
      last_edited: g('xf_ht')!=='' ? 'ht':'ttc',
      tip:g('xf_tip')||null, currency:g('xf_currency')||'CAD',
      amount_currency: document.getElementById('xf_curcheck').checked ? (g('xf_curamount')||null) : null,
    };
    await window.pywebview.api.save_expense(payload);
    editingExpense=null;
    document.getElementById('btnExpense').textContent='Enregistrer';
    document.getElementById('expenseFormWrap').classList.add('hide');
    loadExpenses();
  }

  function fillExpenseForm(e, opts){
    // remplit le formulaire a partir d'une operation existante.
    // opts.keepDate=false -> date du jour (cas duplication)
    openExpenseForm(e.op_type);
    document.getElementById('xf_date').value =
      (opts && opts.keepDate) ? e.date : (curDay||new Date().toISOString().slice(0,10));
    document.getElementById('xf_label').value=e.label||'';
    document.getElementById('xf_third').value=e.third_party||'';
    updateOpTypeUI();
    if(e.category_id) document.getElementById('xf_category').value=e.category_id;
    if(e.tax_code_id) document.getElementById('xf_taxcode').value=e.tax_code_id;
    document.getElementById('xf_notax').checked=e.no_tax;
    document.getElementById('xf_personal').checked=e.is_personal;
    document.getElementById('xf_ht').value=e.amount_ht||'';
    document.getElementById('xf_ttc').value=e.amount_ttc||'';
    const vals=Object.values(e.tax_detail||{});
    document.getElementById('xf_tps').value=vals[0]!==undefined?vals[0]:'';
    document.getElementById('xf_tvq').value=vals[1]!==undefined?vals[1]:'';
    document.getElementById('xf_tip').value=e.tip||'';
    document.getElementById('xf_currency').value=e.currency||'CAD';
    if(e.amount_currency){ document.getElementById('xf_curcheck').checked=true;
      document.getElementById('xf_curamount').classList.remove('hide');
      document.getElementById('xf_curamount').value=e.amount_currency; }
    updateTaxLabels();
  }

  async function editExpense(id){
    const e=await window.pywebview.api.get_expense(id); if(!e) return;
    fillExpenseForm(e, {keepDate:true});
    editingExpense=id;                    // mode modification
    document.getElementById('btnExpense').textContent='Enregistrer';
  }

  async function duplicateExpense(id){
    const e=await window.pywebview.api.get_expense(id); if(!e) return;
    fillExpenseForm(e, {keepDate:false}); // date remise a aujourd'hui
    editingExpense=null;                  // mode CREATION : l'originale reste intacte
    document.getElementById('btnExpense').textContent='Créer la copie';
    document.getElementById('xf_label').focus();
  }
  async function delExpense(id){ await window.pywebview.api.delete_expense(id); loadExpenses(); }

  // --- Catégories de dépenses ---
  let editingExCat=null;
  async function loadExCats(){
    const cats=await window.pywebview.api.list_expense_categories();
    const el=document.getElementById('ec_list');
    el.innerHTML = !cats.length ? '<div class="muted">Aucune catégorie.</div>' :
      cats.map(c=>`<div class="row"><span>${esc(c.name)}</span>
        <span><span class="iconbtn" onclick="renameExCat(${c.id},'${esc(c.name).replace(/'/g,"\\'")}')">✎</span>
        <span class="iconbtn" onclick="delExCat(${c.id})">🗑</span></span></div>`).join('');
  }
  async function addExCat(){
    const n=document.getElementById('ec_name').value; if(!n.trim()) return;
    await window.pywebview.api.add_expense_category(n);
    document.getElementById('ec_name').value=''; loadExCats();
  }
  async function renameExCat(id,old){
    const n=prompt('Renommer la catégorie :', old); if(n===null||!n.trim()) return;
    await window.pywebview.api.update_expense_category(id,n); loadExCats();
  }
  async function delExCat(id){
    if(!confirm('Supprimer cette catégorie ?')) return;
    await window.pywebview.api.delete_expense_category(id); loadExCats();
  }

  // --- Codes de taxe ---
  let editingTaxCode=null;
  async function loadTaxes(){
    const codes=await window.pywebview.api.list_tax_codes();
    const el=document.getElementById('tax_list');
    el.innerHTML = !codes.length ? '<div class="muted">Aucun code.</div>' :
      codes.map(c=>{
        const lines=c.lines.map(l=>`${esc(l.name)} ${l.rate}%`).join(' + ')||'aucune taxe';
        return `<div class="entry"><div class="row">
          <span><b>${esc(c.name)}</b>${c.is_default?' <span class="tag">défaut</span>':''}</span>
          <span><span class="iconbtn" onclick="editTaxCode(${c.id})">✎</span>
          <span class="iconbtn" onclick="delTaxCode(${c.id})">🗑</span></span></div>
          <div class="muted" style="font-size:12px">${lines}</div></div>`;
      }).join('');
    if(editingTaxCode===null) resetTaxForm();
  }
  function addTaxLineRow(name='', rate=''){
    const wrap=document.getElementById('tx_lines');
    const div=document.createElement('div');
    div.className='form'; div.style.margin='0 0 6px';
    div.innerHTML=`<input type="text" placeholder="nom (ex: TPS)" class="txln-name" value="${name?esc(name):''}" style="width:140px">
      <input type="number" step="any" placeholder="taux %" class="txln-rate" value="${rate!==''&&rate!==null?rate:''}" style="width:100px">
      <span class="iconbtn" onclick="this.closest('.form').remove()">🗑</span>`;
    wrap.appendChild(div);
  }
  function resetTaxForm(){
    editingTaxCode=null;
    document.getElementById('tax_form_title').textContent='Nouveau code de taxe';
    document.getElementById('tx_name').value='';
    document.getElementById('tx_default').checked=false;
    document.getElementById('tx_lines').innerHTML='';
    addTaxLineRow();
    document.getElementById('btnTax').textContent='Enregistrer';
  }
  async function editTaxCode(id){
    const c=await window.pywebview.api.get_tax_code(id); if(!c) return;
    editingTaxCode=id;
    document.getElementById('tax_form_title').textContent='Modifier : '+c.name;
    document.getElementById('tx_name').value=c.name;
    document.getElementById('tx_default').checked=!!c.is_default;
    document.getElementById('tx_lines').innerHTML='';
    if(c.lines.length) c.lines.forEach(l=>addTaxLineRow(l.name,l.rate));
    else addTaxLineRow();
    document.getElementById('btnTax').textContent='Enregistrer';
  }
  function collectTaxLines(){
    const lines=[];
    document.querySelectorAll('#tx_lines .form').forEach(row=>{
      const name=row.querySelector('.txln-name').value.trim();
      const rate=row.querySelector('.txln-rate').value;
      if(name && rate!=='') lines.push({name, rate:parseFloat(rate)});
    });
    return lines;
  }
  async function submitTaxCode(){
    const name=document.getElementById('tx_name').value;
    if(!name.trim()){ alert('Donne un nom au code.'); return; }
    const lines=collectTaxLines();
    const isDef=document.getElementById('tx_default').checked;
    if(editingTaxCode) await window.pywebview.api.update_tax_code(editingTaxCode,name,lines,isDef);
    else await window.pywebview.api.add_tax_code(name,lines,isDef);
    editingTaxCode=null;
    loadTaxes();
  }
  async function delTaxCode(id){
    if(!confirm('Supprimer ce code de taxe ?')) return;
    await window.pywebview.api.delete_tax_code(id);
    if(editingTaxCode===id) editingTaxCode=null;
    loadTaxes();
  }

  const FREQ_FR={daily:'quotidien',weekly:'hebdo',monthly:'mensuel'};
  function updateRecurType(){
    const isBudget=document.getElementById('rf_type').value==='budget';
    document.getElementById('rf_budget_fields').classList.toggle('hide', !isBudget);
  }
  async function openRecurForm(){
    // remplit le menu des comptes et pre-remplit la date de debut
    const accs=await window.pywebview.api.recurrence_accounts();
    const sel=document.getElementById('rf_account');
    sel.innerHTML=accs.map(a=>`<option value="${a.id}">${esc(a.name)}</option>`).join('');
    if(!document.getElementById('rf_start').value && curDay)
      document.getElementById('rf_start').value=curDay;
    updateRecurType();
    document.getElementById('recurForm').classList.remove('hide');
  }
  async function generateAhead(){
    const until=document.getElementById('gen_until').value;
    if(!until){ alert('Choisis une date limite pour la génération.'); return; }
    const r=await window.pywebview.api.generate_recurrences_until(until);
    document.getElementById('gen_info').textContent=
      `${r.count} occurrence(s) générée(s) jusqu'au ${r.until}.`;
    renderRecur(); // rafraîchit la liste + la vue du jour
  }
  async function submitRecur(){
    const title=document.getElementById('rf_title').value;
    if(!title.trim()){ return; }
    const type=document.getElementById('rf_type').value;
    const freq=document.getElementById('rf_freq').value;
    const interval=document.getElementById('rf_interval').value||1;
    const start=document.getElementById('rf_start').value;
    const tags=document.getElementById('rf_tags').value;
    let account=null, amount=null, isExpense=false;
    if(type==='budget'){
      account=document.getElementById('rf_account').value||null;
      amount=document.getElementById('rf_amount').value||null;
      isExpense=document.getElementById('rf_expense').checked;
      if(!account){ alert('Choisis un compte (ou crée-en un dans le Budget).'); return; }
      if(amount===''||amount===null){ alert('Indique un montant.'); return; }
    }
    renderRecur(await window.pywebview.api.add_recurrence(
      title, freq, interval, start||null, tags||null, type, account, amount, isExpense));
    ['rf_title','rf_tags','rf_start','rf_amount'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('rf_interval').value='1';
    document.getElementById('rf_freq').value='weekly';
    document.getElementById('rf_type').value='task';
    document.getElementById('rf_expense').checked=true;
    updateRecurType();
    document.getElementById('recurForm').classList.add('hide');
  }
  async function loadRecur(){
    const recs=await window.pywebview.api.load_recurrences();
    const tb=document.getElementById('r_list');
    if(!recs.length){ tb.innerHTML='<tr><td colspan="9" class="muted">Aucune récurrence.</td></tr>'; return; }
    tb.innerHTML=recs.map(r=>{
      let f=FREQ_FR[r.freq]||r.freq;
      if(r.interval>1) f='tous les '+r.interval+' ('+f+')';
      let typeCell, amountCell;
      if(r.type==='budget'){
        typeCell='<span class="tag">budget</span>';
        const amt=(r.amount!=null)?r.amount.toFixed(2):'';
        const cls=(r.amount!=null&&r.amount<0)?'neg':'pos';
        amountCell=`<span class="${cls}">${amt}</span>`+
          (r.account?` <span class="muted" style="font-size:12px">${esc(r.account)}</span>`:'');
      } else {
        typeCell='<span class="muted">tâche</span>';
        amountCell='<span class="muted">—</span>';
      }
      return `<tr>
        <td>${esc(r.title)}</td><td>${typeCell}</td>
        <td style="text-align:right">${amountCell}</td>
        <td>${f}</td><td>${r.start}</td>
        <td>${r.last_run||'<span class="muted">jamais</span>'}</td>
        <td class="tag">${r.tags.map(t=>'#'+esc(t)).join(' ')}</td>
        <td class="${r.active?'pos':'muted'}">${r.active?'actif':'inactif'}</td>
        <td style="text-align:right">
          <button class="small" onclick="togRecur(${r.id},${!r.active})">${r.active?'désactiver':'activer'}</button>
          <button class="small danger" onclick="delRecur(${r.id})">🗑</button>
        </td></tr>`;
    }).join('');
  }
  async function togRecur(id,act){ renderRecur(await window.pywebview.api.toggle_recurrence(id,act)); }
  async function delRecur(id){ renderRecur(await window.pywebview.api.delete_recurrence(id)); }
  function renderRecur(){ loadRecur(); reload(); }

  function render(d){
    curDay = d.day.iso;
    refreshBadges();
    const navd=document.getElementById('nav_date'); if(navd) navd.value=d.day.iso;
    document.getElementById('dateLabel').textContent =
      d.day.human + (d.day.is_today ? "  (aujourd'hui)" : "");

    const t=document.getElementById('tasks');
    t.innerHTML = !d.tasks.length ? '<div class="muted">Aucune tâche.</div>' :
      d.tasks.map(x=>`<div class="task ${x.done?'done':''}">
        <span class="pdot ${x.priority?'on':''}" title="${x.priority?'prioritaire':'normale'} (cliquer pour basculer)" onclick="togglePriority(${x.id}, ${x.priority?0:1})"></span>
        <input type="checkbox" ${x.done?'checked':''} onchange="toggle(${x.id},this)">
        <span class="t">${esc(x.title)}</span>
        ${x.tags.map(tg=>`<span class="tag">#${esc(tg)}</span>`).join(' ')}
        <span style="flex:1"></span>
        <span class="iconbtn" onclick="editTask(${x.id})">✎</span>
        <span class="iconbtn" onclick="delTask(${x.id})">🗑</span></div>`).join('');

    const j=document.getElementById('journal');
    j.innerHTML = !d.journal.length ? '<div class="muted">Aucune entrée.</div>' :
      d.journal.map(e=>{ let h='<div class="entry"><div class="row"><div>';
        if(e.mood!==null) h+=`<span class="pos">Humeur : ${e.mood}/10</span>`;
        const m=Object.entries(e.metrics).map(([k,v])=>`${k} ${v}`).join('   ');
        if(m) h+=` <span style="color:var(--cyan)">${esc(m)}</span>`;
        h+=`</div><span><span class="iconbtn" onclick="editJournal(${e.id})">✎</span>`;
        h+=`<span class="iconbtn" onclick="delJournal(${e.id})">🗑</span></span></div>`;
        h+=`<div>${esc(e.body)}</div>`;
        if(e.tags.length) h+=`<div class="tag">${e.tags.map(x=>'#'+esc(x)).join(' ')}</div>`;
        return h+'</div>'; }).join('');

    const s=document.getElementById('sport');
    s.innerHTML = !d.sport.length ? '<div class="muted">Aucune activité.</div>' :
      d.sport.map((a,i)=>{ const p=[a.activity];
        if(a.duration)p.push(a.duration+' min'); if(a.distance)p.push(a.distance+' km');
        if(a.pace)p.push(a.pace+'/km'); if(a.calories)p.push(a.calories+' kcal');
        return `<div class="row"><span>${esc(p.join('  ·  '))}</span><span>
          <span class="iconbtn" onclick="editSport(${a.id})">✎</span>
          <span class="iconbtn" onclick="delSport(${a.id})">🗑</span></span></div>`; }).join('');

    const f=document.getElementById('food'); let fh='';
    if(!d.food.length){ fh='<div class="muted">Rien de consommé.</div>'; }
    else { fh=d.food.map(x=>`<div class="row"><span>${esc(x.label)} <span class="muted">${esc(x.qty)}</span></span><span>${x.kcal} kcal <span class="iconbtn" onclick="editFood(${x.id})">✎</span><span class="iconbtn" onclick="delFood(${x.id})">🗑</span></span></div>`).join('');
      const tot=d.food_totals; fh+='<div class="sep"></div>';
      fh+=`<div class="total">${tot.kcal} kcal &nbsp; P ${tot.protein}g &nbsp; G ${tot.carbs}g &nbsp; F ${tot.fiber}g</div>`;
      fh+=`<div class="muted">Lipides — saturés ${tot.fat_sat}g · insaturés ${tot.fat_unsat}g</div>`; }
    f.innerHTML=fh;

    const ac=document.getElementById('accounts');
    ac.innerHTML = !d.accounts.length ? '<div class="muted">Aucun compte.</div>' :
      d.accounts.map(a=>`<div class="row"><span>${esc(a.name)}</span><span class="${a.balance>=0?'pos':'neg'}">${a.balance.toFixed(2)}</span></div>`).join('')
      + '<div class="sep"></div>'
      + `<div class="row"><span><b>Total</b></span><span class="${d.accounts_total>=0?'pos':'neg'}"><b>${d.accounts_total.toFixed(2)}</b></span></div>`;
  }

  let budgetAccounts=[], editingOp=null, budgetFilter=null;

  function fillAccountSelect(id, accs){
    const s=document.getElementById(id);
    const keepAll = id==='flt_account';
    s.innerHTML=(keepAll?'<option value="">tous comptes</option>':'')
      + accs.map(a=>`<option value="${a.id}">${esc(a.name)}</option>`).join('');
  }

  async function loadBudget(){
    const flt = budgetFilter || {};
    const b=await window.pywebview.api.load_budget(
      flt.since||null, flt.until||null, flt.account_id||null, flt.tag||null,
      budgetFilter===null);
    budgetAccounts=b.accounts;
    ['of_account','tr_from','tr_to','flt_account'].forEach(id=>fillAccountSelect(id,b.accounts));
    if(budgetFilter && budgetFilter.account_id) document.getElementById('flt_account').value=budgetFilter.account_id;

    // info période affichée
    const ap=b.applied;
    let info;
    if(ap.default) info = `Vue par défaut : mois en cours + 10 opérations à venir`;
    else info = `Filtré : ${ap.since||'…'} → ${ap.until||'…'}`;
    info += `  ·  ${b.operations.length} opération(s)  ·  somme ${b.period_sum>=0?'+':''}${b.period_sum.toFixed(2)}`;
    document.getElementById('flt_info').textContent = info;

    document.getElementById('b_ops').innerHTML = !b.operations.length ?
      '<tr><td colspan="6" class="muted">Aucune opération.</td></tr>' :
      b.operations.map(o=>`<tr><td>${o.date}</td><td>${esc(o.title)}</td>
        <td class="tag">${o.tags.map(t=>'#'+esc(t)).join(' ')}</td>
        <td style="text-align:right" class="${o.amount>=0?'pos':'neg'}">${o.amount>=0?'+':''}${o.amount.toFixed(2)}</td>
        <td style="text-align:center"><input type="checkbox" ${o.reconciled?'checked':''} onchange="toggleReconciled(${o.id}, this)" title="Rapproché avec le compte bancaire"></td>
        <td style="text-align:right"><span class="iconbtn" onclick="editOp(${o.id})">✎</span><span class="iconbtn" onclick="delOp(${o.id})">🗑</span></td></tr>`).join('');
    document.getElementById('b_accounts').innerHTML =
      b.accounts.map(a=>`<div class="row"><span>${esc(a.name)}</span><span><span class="${a.balance>=0?'pos':'neg'}">${a.balance.toFixed(2)}</span> <span class="iconbtn" onclick="delAccount(${a.id},'${esc(a.name)}')">🗑</span></span></div>`).join('')
      + '<div class="sep"></div>'
      + `<div class="row"><span><b>Total</b></span><span class="${b.accounts_total>=0?'pos':'neg'}"><b>${b.accounts_total.toFixed(2)}</b></span></div>`;
    if(!b.categories.length){
      document.getElementById('b_cats').innerHTML='<div class="muted">Aucune donnée.</div>';
    } else {
      const catTotal=b.categories.reduce((s,c)=>s+c.total,0);
      document.getElementById('b_cats').innerHTML=
        b.categories.map(c=>`<div class="row"><span>${esc(c.category)} <span class="muted">(${c.count})</span></span><span class="${c.total>=0?'pos':'neg'}">${c.total>=0?'+':''}${c.total.toFixed(2)}</span></div>`).join('')
        + '<div class="sep"></div>'
        + `<div class="row total"><span>Total (revenus − dépenses)</span><span class="${catTotal>=0?'pos':'neg'}">${catTotal>=0?'+':''}${catTotal.toFixed(2)}</span></div>`;
    }
  }

  function applyFilter(){
    budgetFilter={
      since:document.getElementById('flt_since').value,
      until:document.getElementById('flt_until').value,
      account_id:document.getElementById('flt_account').value,
      tag:document.getElementById('flt_tag').value,
    };
    loadBudget();
  }
  function resetFilter(){
    budgetFilter=null;
    ['flt_since','flt_until','flt_tag'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('flt_account').value='';
    loadBudget();
  }

  async function submitAccount(){
    const name=document.getElementById('ac_name').value;
    if(!name.trim()) return;
    await window.pywebview.api.add_account(name,
      document.getElementById('ac_balance').value||null,
      document.getElementById('ac_opened').value||null);
    ['ac_name','ac_balance','ac_opened'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('accForm').classList.add('hide');
    loadBudget();
  }
  async function delAccount(id,name){
    if(!confirm('Supprimer le compte « '+name+' » et toutes ses opérations ?')) return;
    await window.pywebview.api.delete_account(id); loadBudget();
  }

  async function submitOp(){
    const acc=document.getElementById('of_account').value;
    const title=document.getElementById('of_title').value;
    const amount=document.getElementById('of_amount').value;
    const date=document.getElementById('of_date').value;
    const tags=document.getElementById('of_tags').value;
    const exp=document.getElementById('of_expense').checked;
    if(amount==='') return;
    if(editingOp){
      await window.pywebview.api.update_operation(editingOp,amount,title,date||null,acc,tags,exp);
      editingOp=null; document.getElementById('btnOp').textContent='Ajouter';
    } else {
      await window.pywebview.api.add_operation(acc,amount,title,date||null,tags,exp);
    }
    ['of_title','of_amount','of_date','of_tags'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('of_expense').checked=false;
    document.getElementById('opForm').classList.add('hide');
    loadBudget();
  }
  async function editOp(id){
    const o=await window.pywebview.api.get_operation(id); if(!o) return;
    editingOp=id;
    document.getElementById('of_title').value=o.title;
    document.getElementById('of_amount').value=Math.abs(o.amount);
    document.getElementById('of_date').value=o.date;
    document.getElementById('of_tags').value=o.tags;
    document.getElementById('of_expense').checked=o.amount<0;
    if(o.account_id) document.getElementById('of_account').value=o.account_id;
    document.getElementById('btnOp').textContent= o.is_transfer ? 'Enregistrer (virement)' : 'Enregistrer';
    document.getElementById('opForm').classList.remove('hide');
  }
  async function delOp(id){ await window.pywebview.api.delete_operation(id); loadBudget(); }
  async function toggleReconciled(id, el){
    // enregistre le rapprochement sans passer par le formulaire de modification
    const r=await window.pywebview.api.set_reconciled(id, el.checked);
    if(!r || !r.ok){ el.checked=!el.checked; return; } // annule visuellement si échec
  }

  async function submitTransfer(){
    const from=document.getElementById('tr_from').value;
    const to=document.getElementById('tr_to').value;
    const amount=document.getElementById('tr_amount').value;
    const date=document.getElementById('tr_date').value;
    if(amount===''||from===to) { if(from===to) alert('Choisis deux comptes différents.'); return; }
    await window.pywebview.api.add_transfer(from,to,amount,null,date||null,null);
    document.getElementById('tr_amount').value=''; document.getElementById('tr_date').value='';
    document.getElementById('transferForm').classList.add('hide');
    loadBudget();
  }

  window.addEventListener('pywebviewready', reload);
</script>
</body>
</html>
"""


def run() -> None:
    import os

    db.init_db()
    api = Api()
    api.window = webview.create_window("Task365", html=HTML,
                                       js_api=api, width=920, height=720)

    icon_path = (os.environ.get("TASK365_ICON")
                 or os.environ.get("FLEURDELIX_ICON")  # ancien nom, compat
                 or os.path.join(os.path.dirname(__file__), "icon.png"))
    start_kwargs = {}
    if os.path.isfile(icon_path):
        start_kwargs["icon"] = icon_path
    try:
        webview.start(**start_kwargs)
    except TypeError:
        webview.start()


if __name__ == "__main__":
    run()
